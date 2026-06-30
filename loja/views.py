# ---------------------------------------------------------------------------
# Página de gestão da newsletter
# ---------------------------------------------------------------------------
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required(login_url='/gestao/login/')
def newsletter_gestao(request):
    inscritos = NewsletterInscricao.objects.all().order_by('-data_inscricao')
    inscritos_ativos = inscritos.filter(ativo=True).count()
    return render(request, 'gestao/newsletter_gestao.html', {
        'active_page': 'newsletter',
        'inscritos': inscritos,
        'inscritos_ativos': inscritos_ativos,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_enviar_newsletter(request):
    """Envia um e-mail em grupo para todos os inscritos ativos na newsletter."""
    if request.method != 'POST':
        return redirect('newsletter_gestao')

    assunto = request.POST.get('assunto', '').strip()
    mensagem = request.POST.get('mensagem', '').strip()

    if not assunto or not mensagem:
        messages.error(request, 'Preencha o assunto e a mensagem antes de enviar.')
        return redirect('newsletter_gestao')

    destinatarios = list(
        NewsletterInscricao.objects.filter(ativo=True).values_list('email', flat=True)
    )

    if not destinatarios:
        messages.warning(request, 'Não existem inscritos ativos para enviar o e-mail.')
        return redirect('newsletter_gestao')

    try:
        enviados = enviar_email_newsletter(assunto, mensagem, destinatarios)
        messages.success(
            request,
            f'E-mail enviado com sucesso para {len(destinatarios)} inscrito(s).'
        )
    except Exception as e:
        messages.error(request, f'Erro ao enviar e-mail: {e}')

    return redirect('newsletter_gestao')
# ---------------------------------------------------------------------------
# Utilitário para envio de e-mails em grupo para inscritos na newsletter
# ---------------------------------------------------------------------------
from django.core.mail import send_mail
from django.conf import settings

def enviar_email_newsletter(assunto, mensagem, destinatarios=None):
    """
    Envia e-mail para todos os inscritos ativos na newsletter ou para destinatários específicos.
    destinatarios: lista de e-mails (opcional)
    """
    if destinatarios is None:
        destinatarios = list(NewsletterInscricao.objects.filter(ativo=True).values_list('email', flat=True))
    if not destinatarios:
        return 0
    return send_mail(
        assunto,
        mensagem,
        settings.DEFAULT_FROM_EMAIL,
        destinatarios,
        fail_silently=False,
    )
import logging
import re
import json
from django.db import models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from django.db.models import Sum, Count, Q
from .models import Encomenda, ItemEncomenda, Produto, Cliente, Funcionario, Categoria, MovimentoCaixa, VisitaSite, NewsletterInscricao, MovimentoStock, Promocao, PagamentoSalario, LoteImportacao, ItemLoteImportacao, Banner, ImagemGaleria, Cupao
from .forms import NewsletterInscricaoForm

logger = logging.getLogger(__name__)
# ---------------------------------------------------------------------------
# Newsletter - Cadastro de e-mails
# ---------------------------------------------------------------------------
from django.views.decorators.http import require_POST

@require_POST
def newsletter_inscrever(request):
    form = NewsletterInscricaoForm(request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({'success': True, 'message': 'Inscrição realizada com sucesso!'}, status=201)
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)
from django.contrib.auth.models import User, Group
from .utils.auth import is_operador
from .utils.validators import limpar_bi, limpar_telefone, BI_REGEX, TELEFONE_REGEX
from .utils.stock import reverter_stock_encomenda
from .utils.promotions import get_promocoes_activas
from .utils.encomenda import contar_por_status
from .utils.slug import gerar_slug_unico


# ---------------------------------------------------------------------------
# Controlo de acesso por nível
# ---------------------------------------------------------------------------

def _is_operador(user):
    """Verifica se o utilizador pertence ao grupo Operador (acesso restrito)."""
    return is_operador(user)


def _block_operador(view_func):
    """Decorador que bloqueia o acesso de operadores a uma view de gestão."""
    from functools import wraps

    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if _is_operador(request.user):
            messages.error(request, 'Sem permissão para aceder a esta área.')
            return redirect('gestao_dashboard')
        return view_func(request, *args, **kwargs)

    return _wrapped


def _registar_venda_no_caixa(encomenda, user=None):
    """Cria automaticamente uma entrada de caixa para uma venda finalizada.

    Idempotente: não cria duplicados se já existir um movimento associado.
    """
    if encomenda.status != 'finalizada':
        return None
    # Backfill snapshot do pre\u00e7o de compra para itens que ainda n\u00e3o o t\u00eam
    for item in encomenda.itens.select_related('produto').all():
        if (not item.preco_custo_unitario or float(item.preco_custo_unitario) == 0) and item.produto:
            custo = float(item.produto.preco_compra or 0)
            if custo > 0:
                item.preco_custo_unitario = custo
                item.save(update_fields=['preco_custo_unitario'])
    if MovimentoCaixa.objects.filter(encomenda=encomenda, tipo='entrada', categoria='venda').exists():
        return None
    total = encomenda.total()
    if not total or float(total) <= 0:
        return None
    return MovimentoCaixa.objects.create(
        tipo='entrada',
        categoria='venda',
        valor=total,
        descricao=f'Venda #{encomenda.pk} — {encomenda.nome_cliente}',
        encomenda=encomenda,
        criado_por=user if user and user.is_authenticated else None,
    )

# ---------------------------------------------------------------------------
# Loja pública
# ---------------------------------------------------------------------------

def home(request):
    # Regista acesso (1 por sessão de 30min para não inflacionar)
    if not request.session.session_key:
        request.session.save()
    sessao = request.session.session_key or ''
    ja_contado = request.session.get('visita_registada')
    if not ja_contado:
        ip = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')
        VisitaSite.objects.create(
            ip=ip or None,
            user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:300],
            referer=(request.META.get('HTTP_REFERER') or '')[:200],
            sessao=sessao,
        )
        request.session['visita_registada'] = True
        request.session.set_expiry(1800)

    total_visitas = VisitaSite.objects.count()
    total_clientes = Cliente.objects.count()

    # Promoções activas (em curso) marcadas para landing/carrossel
    from .models import Promocao
    from datetime import date as _date
    hoje = _date.today()
    activas = [p for p in Promocao.objects.filter(activo=True).prefetch_related('produtos')
               if p.esta_decorrer(hoje)]
    promocoes_carrossel = [p for p in activas if p.mostrar_carrossel]
    promocoes_landing = [p for p in activas if p.mostrar_landing]

    masculinos = list(Produto.objects.filter(disponivel=True, genero='masculino').order_by('-desconto_percentagem', 'nome')[:4])
    femininos  = list(Produto.objects.filter(disponivel=True, genero='feminino').order_by('-desconto_percentagem', 'nome')[:4])
    unissex    = list(Produto.objects.filter(disponivel=True, genero='unissex').order_by('-desconto_percentagem', 'nome')[:4])

    import json as _json
    produtos_busca_json = _json.dumps([
        {'name': p.nome, 'brand': p.marca or '', 'category': p.genero}
        for p in Produto.objects.filter(disponivel=True).only('nome', 'marca', 'genero')
    ], ensure_ascii=False)

    banners_carrossel = list(Banner.objects.filter(activo=True).order_by('ordem'))

    return render(request, 'loja/index.html', {
        'total_visitas': total_visitas,
        'total_clientes': total_clientes,
        'promocoes_carrossel': promocoes_carrossel,
        'promocoes_landing': promocoes_landing,
        'banners_carrossel': banners_carrossel,
        'masculinos': masculinos,
        'femininos': femininos,
        'unissex': unissex,
        'produtos_busca_json': produtos_busca_json,
    })

def colecoes(request):
    _activas, _carrossel, promocoes_publicas = get_promocoes_activas()
    masculinos = Produto.objects.filter(disponivel=True, genero='masculino').order_by('-desconto_percentagem', 'marca', 'nome')
    femininos  = Produto.objects.filter(disponivel=True, genero='feminino').order_by('-desconto_percentagem', 'marca', 'nome')
    unissex    = Produto.objects.filter(disponivel=True, genero='unissex').order_by('-desconto_percentagem', 'marca', 'nome')
    return render(request, 'loja/colecoes.html', {
        'promocoes_publicas': promocoes_publicas,
        'masculinos': masculinos,
        'femininos': femininos,
        'unissex': unissex,
        'total': masculinos.count() + femininos.count() + unissex.count(),
    })


def produto_detalhe(request, pk):
    produto = get_object_or_404(Produto, pk=pk, disponivel=True)
    relacionados = Produto.objects.filter(
        disponivel=True, genero=produto.genero
    ).exclude(pk=pk).order_by('?')[:4]
    return render(request, 'loja/produto_detalhe.html', {
        'produto': produto,
        'relacionados': relacionados,
    })


def promocao_publica(request, slug):
    promo = get_object_or_404(Promocao, slug=slug, activo=True)
    produtos = list(promo.produtos.filter(disponivel=True).order_by('genero', 'nome'))
    desc = promo.desconto_percentagem or 0
    for prod in produtos:
        try:
            prod.preco_promocional = round(float(prod.preco_venda) * (100 - desc) / 100, 2)
        except (TypeError, ValueError):
            prod.preco_promocional = prod.preco_venda
    return render(request, 'loja/promocao.html', {
        'promo': promo,
        'produtos': produtos,
    })

def encomendas(request):
    import json
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        telefone = request.POST.get('telefone', '').strip()
        email = request.POST.get('email', '').strip()
        morada = request.POST.get('morada', '').strip()
        notas = request.POST.get('notas', '').strip()
        itens_json = request.POST.get('itens_json', '[]').strip()

        if not nome or not telefone:
            messages.error(request, 'Nome e telefone são obrigatórios.')
            return render(request, 'loja/encomendas.html', {'form_data': request.POST})

        forma_pag = request.POST.get('forma_pagamento', 'avista')
        if forma_pag not in ('avista', 'parcelado'):
            forma_pag = 'avista'

        # Cupão de desconto
        cupao_codigo = request.POST.get('cupao_codigo', '').strip().upper()
        cupao_desconto = float(request.POST.get('cupao_desconto', '0') or 0)
        cupao_obj = None
        if cupao_codigo and cupao_desconto > 0:
            try:
                cupao_obj = Cupao.objects.get(codigo=cupao_codigo)
                valido, msg = cupao_obj.esta_valido()
                if not valido:
                    cupao_obj = None
                    cupao_desconto = 0
            except Cupao.DoesNotExist:
                cupao_desconto = 0

        from django.db import transaction
        with transaction.atomic():
            encomenda = Encomenda.objects.create(
                nome_cliente=nome,
                telefone=telefone,
                email=email,
                morada=morada,
                notas=notas,
                origem='online',
                forma_pagamento=forma_pag,
                status='em_curso',
            )

            try:
                itens = json.loads(itens_json)
                total_carrinho = sum(float(item.get('price', 0)) * int(item.get('qty', 1)) for item in itens)

                for item in itens:
                    nome_prod = item.get('name', '').strip()
                    preco = float(item.get('price', 0))
                    qty = int(item.get('qty', 1))
                    if nome_prod and qty > 0:
                        produto = Produto.objects.filter(nome=nome_prod, disponivel=True).first()
                        if produto and produto.stock >= qty:
                            produto.stock -= qty
                            produto.save(update_fields=['stock'])
                            MovimentoStock.objects.create(
                                produto=produto,
                                tipo='saida',
                                quantidade=-qty,
                                descricao=f'Encomenda online #{encomenda.pk}',
                                encomenda=encomenda,
                            )
                            ItemEncomenda.objects.create(
                                encomenda=encomenda,
                                produto=produto,
                                nome_produto=nome_prod,
                                preco_unitario=preco,
                                preco_custo_unitario=produto.preco_compra or 0,
                                quantidade=qty,
                            )
                        else:
                            ItemEncomenda.objects.create(
                                encomenda=encomenda,
                                nome_produto=nome_prod,
                                preco_unitario=preco,
                                quantidade=qty,
                            )
            except (json.JSONDecodeError, ValueError, TypeError):
                pass

            # Incrementar uso do cupão
            if cupao_obj:
                cupao_obj.usos += 1
                cupao_obj.save(update_fields=['usos'])
                encomenda.notas = (encomenda.notas or '') + f'\n[Cupão: {cupao_obj.codigo} — Desconto: {cupao_desconto:,.0f} Kz]'
                encomenda.save(update_fields=['notas'])

        request.session['ultima_encomenda_id'] = encomenda.pk
        return redirect('encomenda_sucesso')

    else:
        import json
        carrinho_cookie = request.COOKIES.get('carrinho_checkout', '[]')
        try:
            carrinho = json.loads(carrinho_cookie)
        except (json.JSONDecodeError, ValueError):
            carrinho = []

        return render(request, 'loja/encomendas.html', {'itens': carrinho})

def encomenda_sucesso(request):
    encomenda_id = request.session.pop('ultima_encomenda_id', None)
    return render(request, 'loja/encomenda_sucesso.html', {'encomenda_id': encomenda_id})

def galeria(request):
    imagens = ImagemGaleria.objects.filter(activo=True).order_by('ordem', '-criado_em')
    return render(request, 'loja/galeria.html', {'imagens': imagens})

def fidelidade(request):
    return render(request, 'loja/fidelidade.html')

def contactos(request):
    return render(request, 'loja/contactos.html')


# ---------------------------------------------------------------------------
# Painel de Gestão — Login / Logout
# ---------------------------------------------------------------------------

def gestao_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('gestao_dashboard')

    erro = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        logger.warning(f'[LOGIN] POST recebido: username={username!r}, host={request.get_host()}, path={request.path_info}, method={request.method}')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            logger.warning(f'[LOGIN] Autenticado: user={user.username}, is_staff={user.is_staff}, is_active={user.is_active}')
            login(request, user)
            return redirect('gestao_dashboard')
        else:
            logger.warning(f'[LOGIN] Falhou: user={user}, username={username!r}')
            erro = 'Utilizador ou senha inválidos, ou sem permissão de acesso.'

    return render(request, 'gestao/login.html', {'erro': erro})


@require_POST
def gestao_logout(request):
    logout(request)
    return redirect('gestao_login')


@staff_member_required(login_url='/gestao/login/')
def gestao_guia(request):
    return render(request, 'gestao/guia.html', {'active_page': 'guia'})


# ---------------------------------------------------------------------------
# Painel de Gestão (requer login de staff)
# ---------------------------------------------------------------------------

@staff_member_required(login_url='/gestao/login/')
def gestao_dashboard(request):
    from datetime import timedelta
    import json as _json

    hoje = timezone.now().date()

    # Últimos 14 dias: vendas finalizadas + receita por dia
    dias = [hoje - timedelta(days=i) for i in range(13, -1, -1)]
    vendas_dia_qs = (
        Encomenda.objects
        .filter(status='finalizada', criado_em__date__gte=dias[0])
        .prefetch_related('itens')
    )
    mapa_vendas = {d: 0 for d in dias}
    mapa_receita = {d: 0.0 for d in dias}
    for e in vendas_dia_qs:
        d = e.criado_em.date()
        if d in mapa_vendas:
            mapa_vendas[d] += 1
            mapa_receita[d] += sum(float(i.preco_unitario) * i.quantidade for i in e.itens.all())

    grafico_dias_labels = [d.strftime('%d/%m') for d in dias]
    grafico_vendas_data = [mapa_vendas[d] for d in dias]
    grafico_receita_data = [round(mapa_receita[d], 2) for d in dias]

    # Distribuição por estado (todas as encomendas)
    estados_count = {
        'em_curso': Encomenda.objects.filter(status='em_curso').count(),
        'finalizada': Encomenda.objects.filter(status='finalizada').count(),
        'cancelada': Encomenda.objects.filter(status='cancelada').count(),
    }

    # Top 5 produtos do mês
    inicio_mes = hoje.replace(day=1)
    top_produtos_mes = (
        ItemEncomenda.objects
        .filter(encomenda__status='finalizada', encomenda__criado_em__date__gte=inicio_mes)
        .values('nome_produto')
        .annotate(qty=Sum('quantidade'))
        .order_by('-qty')[:5]
    )
    top_prod_labels = [p['nome_produto'] for p in top_produtos_mes]
    top_prod_data = [p['qty'] for p in top_produtos_mes]

    context = {
        'active_page': 'dashboard',
        'total_encomendas': Encomenda.objects.count(),
        'pendentes': Encomenda.objects.filter(status='em_curso').count(),
        'encomendas_hoje': Encomenda.objects.filter(criado_em__date=hoje).count(),
        'total_produtos': Produto.objects.count(),
        'stock_baixo': Produto.objects.filter(stock__lt=5).count() if Produto.objects.exists() else None,
        'total_clientes': Cliente.objects.count(),
        'ultimas_encomendas': Encomenda.objects.order_by('-criado_em')[:8],
        'produtos_stock_baixo': Produto.objects.filter(stock__lt=5).order_by('stock')[:6],
        'ultimos_movimentos_stock': MovimentoStock.objects.select_related('produto', 'criado_por').order_by('-criado_em')[:15],
        # Chart data (JSON)
        'grafico_dias_labels': _json.dumps(grafico_dias_labels),
        'grafico_vendas_data': _json.dumps(grafico_vendas_data),
        'grafico_receita_data': _json.dumps(grafico_receita_data),
        'estados_labels': _json.dumps(['Em Curso', 'Finalizada', 'Cancelada']),
        'estados_data': _json.dumps([estados_count['em_curso'], estados_count['finalizada'], estados_count['cancelada']]),
        'top_prod_labels': _json.dumps(top_prod_labels),
        'top_prod_data': _json.dumps(top_prod_data),
    }
    return render(request, 'gestao/dashboard.html', context)


@staff_member_required(login_url='/gestao/login/')
def gestao_vendas(request):

    encomendas = Encomenda.objects.prefetch_related('itens').select_related('vendido_por').order_by('-criado_em')

    contadores = contar_por_status(Encomenda.objects.all())

    receita_total = sum(
        sum(item.subtotal() for item in enc.itens.all())
        for enc in Encomenda.objects.filter(status='finalizada').prefetch_related('itens')
    )

    from django.core.paginator import Paginator
    paginator = Paginator(encomendas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'gestao/vendas.html', {
        'active_page': 'vendas',
        'encomendas': page_obj,
        'page_obj': page_obj,
        'contadores': contadores,
        'total_registos': paginator.count,
        'receita_total': receita_total,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_funcionarios(request):
    funcionarios = Funcionario.objects.select_related('utilizador').order_by('nome')
    return render(request, 'gestao/funcionarios.html', {
        'active_page': 'funcionarios',
        'funcionarios': funcionarios,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_produtos(request):
    produtos = Produto.objects.select_related('categoria').order_by('nome')
    categorias = Categoria.objects.all()
    return render(request, 'gestao/produtos.html', {
        'active_page': 'produtos',
        'produtos': produtos,
        'categorias': categorias,
    })




@staff_member_required(login_url='/gestao/login/')
def gestao_registar_pagamento_parcela2(request, pk):
    """Registar o pagamento da 2ª parcela e finalizar a encomenda."""
    encomenda = get_object_or_404(Encomenda, pk=pk)

    if encomenda.forma_pagamento != 'parcelado':
        messages.error(request, 'Esta encomenda não é parcelada.')
        return redirect('gestao_venda_detalhe', pk=pk)
    if encomenda.parcela2_paga:
        messages.warning(request, 'A 2ª parcela já foi registada anteriormente.')
        return redirect('gestao_fatura', pk=pk)
    if encomenda.status == 'cancelada':
        messages.error(request, 'Não é possível registar pagamento de uma encomenda cancelada.')
        return redirect('gestao_vendas')

    if request.method == 'POST':
        encomenda.parcela2_paga = True
        encomenda.status = 'finalizada'
        encomenda.save(update_fields=['parcela2_paga', 'status', 'atualizado_em'])
        _registar_venda_no_caixa(encomenda, request.user)
        messages.success(request, f'Encomenda #{pk} - 2ª parcela registada e venda finalizada.')
        return redirect('gestao_fatura', pk=pk)
    
    return render(request, 'gestao/confirmar_pagamento_parcela2.html', {
        'active_page': 'vendas',
        'encomenda': encomenda,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_fatura(request, pk):
    encomenda = get_object_or_404(Encomenda.objects.prefetch_related('itens__produto'), pk=pk)
    
    # Bloquear acesso a fatura se pagamento à vista e não finalizada
    if encomenda.forma_pagamento == 'avista' and encomenda.status != 'finalizada':
        messages.error(request, 'Fatura não pode ser gerada. Venda à vista deve ser finalizada automaticamente.')
        return redirect('gestao_vendas')
    itens = encomenda.itens.all()
    total = sum(item.subtotal() for item in itens)
    operador = request.user.get_full_name() or request.user.username
    return render(request, 'gestao/fatura.html', {
        'active_page': 'vendas',
        'encomenda': encomenda,
        'itens': itens,
        'total': total,
        'operador': operador,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_fatura_pdf(request, pk):
    import io
    from django.template.loader import render_to_string
    try:
        from xhtml2pdf import pisa
    except ImportError:
        messages.error(request, 'Biblioteca xhtml2pdf não instalada no servidor.')
        return redirect('gestao_fatura', pk=pk)

    encomenda = get_object_or_404(Encomenda.objects.prefetch_related('itens__produto'), pk=pk)
    itens = list(encomenda.itens.select_related('produto').all())
    total = sum(item.subtotal() for item in itens)

    html = render_to_string('gestao/fatura_pdf.html', {
        'encomenda': encomenda,
        'itens': itens,
        'total': total,
    }, request=request)

    buffer = io.BytesIO()
    result = pisa.CreatePDF(html.encode('utf-8'), dest=buffer, encoding='utf-8')

    if result.err:
        logger.error(f'Erro ao gerar PDF fatura {pk}: {result.err}')
        messages.error(request, 'Erro ao gerar o PDF. Tente novamente.')
        return redirect('gestao_fatura', pk=pk)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fatura_{encomenda.pk:05d}.pdf"'
    return response


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_venda_detalhe(request, pk):
    """Exibe os detalhes completos de uma venda/encomenda."""
    encomenda = get_object_or_404(Encomenda.objects.prefetch_related('itens__produto'), pk=pk)
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'finalizar' and encomenda.status == 'em_curso':
            # Se está em curso e não é parcelado, pode finalizar
            encomenda.status = 'finalizada'
            encomenda.save(update_fields=['status', 'atualizado_em'])
            _registar_venda_no_caixa(encomenda, request.user)
            messages.success(request, f'Venda #{encomenda.pk} finalizada com sucesso.')
            return redirect('gestao_venda_detalhe', pk=pk)
        
        elif acao == 'cancelar' and encomenda.status == 'em_curso':
            motivo = request.POST.get('motivo_cancelamento', '').strip()
            if not motivo:
                messages.error(request, 'Motivo do cancelamento é obrigatório.')
            else:
                encomenda.status = 'cancelada'
                encomenda.motivo_cancelamento = motivo
                encomenda.save(update_fields=['status', 'motivo_cancelamento', 'atualizado_em'])
                reverter_stock_encomenda(encomenda, request.user, descricao_prefixo='Cancelamento')
                messages.success(request, f'Venda #{encomenda.pk} cancelada. Stock revertido.')
                return redirect('gestao_vendas')
        
        elif acao == 'eliminar':
            from django.db import transaction
            with transaction.atomic():
                for item in encomenda.itens.all():
                    if item.produto:
                        item.produto.stock += item.quantidade
                        item.produto.save(update_fields=['stock'])
                MovimentoStock.objects.filter(encomenda=encomenda).delete()
                encomenda.itens.all().delete()
                venda_id = encomenda.pk
                encomenda.delete()
            
            messages.success(request, f'Venda #{venda_id} eliminada com sucesso. Stock revertido.')
            return redirect('gestao_vendas')
    
    itens = encomenda.itens.all()
    total = sum(item.subtotal() for item in itens)
    movimentos = encomenda.movimentos_stock.all().order_by('-criado_em')
    
    return render(request, 'gestao/venda_detalhe.html', {
        'active_page': 'vendas',
        'encomenda': encomenda,
        'itens': itens,
        'total': total,
        'movimentos': movimentos,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_venda_nova(request):
    produtos = Produto.objects.filter(disponivel=True).order_by('nome')
    erros = {}

    if request.method == 'POST':
        tipo_cliente = request.POST.get('tipo_cliente', 'particular')
        nome_empresa = request.POST.get('nome_empresa', '').strip()
        nif          = request.POST.get('nif', '').strip()
        nome_cliente = request.POST.get('nome_cliente', '').strip() or 'Cliente Balcão'
        telefone     = request.POST.get('telefone', '').strip()
        email        = request.POST.get('email', '').strip()
        morada       = request.POST.get('morada', '').strip()
        notas        = request.POST.get('notas', '').strip()
        produto_ids  = request.POST.getlist('produto_id')
        quantidades  = request.POST.getlist('quantidade')
        precos       = request.POST.getlist('preco_unitario')

        # Filtrar linhas vazias
        itens_validos = []
        for pid, qty_raw, preco_raw in zip(produto_ids, quantidades, precos):
            if not pid:
                continue
            try:
                qty = int(qty_raw)
                preco = float(preco_raw.replace(',', '.'))
            except (ValueError, AttributeError):
                erros['itens'] = 'Verifique as quantidades e preços introduzidos.'
                break
            if qty < 1:
                erros['itens'] = 'A quantidade deve ser no mínimo 1.'
                break
            itens_validos.append({'pid': pid, 'qty': qty, 'preco': preco})

        if not itens_validos and 'itens' not in erros:
            erros['itens'] = 'Adicione pelo menos um produto à venda.'

        # Validar stock
        if not erros:
            for item in itens_validos:
                try:
                    p = Produto.objects.get(pk=item['pid'])
                except Produto.DoesNotExist:
                    erros['itens'] = f'Produto inválido seleccionado.'
                    break
                if p.stock < item['qty']:
                    erros['itens'] = f'Stock insuficiente para "{p.nome}" (disponível: {p.stock}).'
                    break
                item['produto'] = p

        # Pagamento parcelado
        forma_pagamento = request.POST.get('forma_pagamento', 'avista')
        valor_parcela1 = request.POST.get('valor_parcela1')
        valor_parcela2 = request.POST.get('valor_parcela2')
        total_venda = sum(item['preco'] * item['qty'] for item in itens_validos)
        if forma_pagamento == 'parcelado':
            try:
                valor_parcela1 = float(valor_parcela1 or 0)
                valor_parcela2 = float(valor_parcela2 or 0)
            except (ValueError, TypeError):
                erros['pagamento'] = 'Valores das parcelas inválidos.'
            if valor_parcela1 < total_venda * 0.5:
                erros['pagamento'] = 'A 1ª parcela deve ser pelo menos 50% do total.'
            if abs((valor_parcela1 + valor_parcela2) - total_venda) > 0.01:
                erros['pagamento'] = 'A soma das parcelas deve ser igual ao total da venda.'
        else:
            valor_parcela1 = None
            valor_parcela2 = None

        if not erros:
            from django.db import transaction
            with transaction.atomic():
                # Determinar status e parcelas
                if forma_pagamento == 'parcelado':
                    status = 'em_curso'
                    parcela1_paga = True  # 1ª parcela paga na hora
                    parcela2_paga = False
                    from datetime import timedelta, date as date_type
                    data_str = request.POST.get('data_proxima_parcela', '').strip()
                    try:
                        data_proxima_parcela = date_type.fromisoformat(data_str) if data_str else timezone.now().date() + timedelta(days=30)
                    except ValueError:
                        data_proxima_parcela = timezone.now().date() + timedelta(days=30)
                else:
                    status = 'finalizada'
                    parcela1_paga = False
                    parcela2_paga = False
                    data_proxima_parcela = None

                encomenda = Encomenda.objects.create(
                    tipo_cliente=tipo_cliente,
                    nome_empresa=nome_empresa,
                    nif=nif,
                    nome_cliente=nome_cliente,
                    telefone=telefone or '—',
                    email=email,
                    morada=morada,
                    notas=notas,
                    status=status,
                    origem='balcao',
                    vendido_por=request.user if request.user.is_authenticated else None,
                    forma_pagamento=forma_pagamento,
                    valor_parcela1=valor_parcela1,
                    valor_parcela2=valor_parcela2,
                    parcela1_paga=parcela1_paga,
                    parcela2_paga=parcela2_paga,
                    data_proxima_parcela=data_proxima_parcela,
                )
                for item in itens_validos:
                    p = item['produto']
                    ItemEncomenda.objects.create(
                        encomenda=encomenda,
                        produto=p,
                        nome_produto=p.nome,
                        preco_unitario=item['preco'],
                        preco_custo_unitario=p.preco_compra or 0,
                        quantidade=item['qty'],
                    )
                    p.stock -= item['qty']
                    p.save(update_fields=['stock'])
                    MovimentoStock.objects.create(
                        produto=p,
                        tipo='saida',
                        quantidade=-item['qty'],
                        descricao=f'Venda: Encomenda #{encomenda.pk}',
                        encomenda=encomenda,
                        criado_por=request.user if request.user.is_authenticated else None,
                    )
                _registar_venda_no_caixa(encomenda, request.user)
            if forma_pagamento == 'parcelado':
                messages.success(request, f'Venda parcelada #{encomenda.pk} criada. 1ª parcela paga. Próximo pagamento: {encomenda.data_proxima_parcela.strftime("%d/%m/%Y")}')
            else:
                messages.success(request, f'Venda #{encomenda.pk} finalizada com sucesso. Stock actualizado.')
            return redirect('gestao_fatura', pk=encomenda.pk)

    import json as _json
    produtos_json = _json.dumps({
        str(p.pk): {
            'nome': p.nome,
            'marca': p.marca,
            'preco': float(p.preco),
            'stock': p.stock,
            'codigo_barras': p.codigo_barras or '',
        }
        for p in produtos
    })

    return render(request, 'gestao/venda_form.html', {
        'active_page': 'vendas',
        'produtos': produtos,
        'produtos_json': produtos_json,
        'erros': erros,
        'post': request.POST,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_cancelar_encomenda(request, pk):
    """Cancelar uma encomenda finalizada e reverter stock."""
    encomenda = get_object_or_404(Encomenda, pk=pk)
    
    if encomenda.status != 'finalizada':
        messages.error(request, 'Apenas encomendas finalizadas podem ser canceladas.')
        return redirect('gestao_vendas')
    
    if request.method == 'POST':
        encomenda.status = 'cancelada'
        encomenda.save(update_fields=['status', 'atualizado_em'])
        reverter_stock_encomenda(encomenda, request.user, descricao_prefixo='Devolução')
        messages.success(request, f'Encomenda #{pk} cancelada com sucesso. Stock devolvido.')
        return redirect('gestao_vendas')
    
    return render(request, 'gestao/confirmar_cancelamento.html', {
        'active_page': 'vendas',
        'encomenda': encomenda,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_funcionario_criar(request, pk=None):
    instancia = get_object_or_404(Funcionario, pk=pk) if pk else None
    utilizadores_disponiveis = User.objects.filter(
        Q(funcionario__isnull=True) | Q(pk=instancia.utilizador_id if instancia and instancia.utilizador_id else 0)
    ).distinct().order_by('username')
    erros = {}

    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        bi = request.POST.get('bi', '').strip()
        telefone = request.POST.get('telefone', '').strip()
        email = request.POST.get('email', '').strip()
        cargo = request.POST.get('cargo', 'operador_caixa')
        turno = request.POST.get('turno', 'integral')
        salario_raw = request.POST.get('salario', '').strip()
        percentagem_comissao_raw = request.POST.get('percentagem_comissao', '2').strip()
        subsidio_alimentacao_raw = request.POST.get('subsidio_alimentacao', '0').strip()
        subsidio_transporte_raw = request.POST.get('subsidio_transporte', '0').strip()
        banco = request.POST.get('banco', '').strip()
        nib_raw = request.POST.get('nib', '').strip().replace(' ', '').replace('-', '')
        iban_raw = request.POST.get('iban', '').strip().replace(' ', '').upper()
        data_admissao = request.POST.get('data_admissao', '').strip()
        activo = request.POST.get('activo') == 'on'
        utilizador_id = request.POST.get('utilizador', '').strip()
        data_nascimento = request.POST.get('data_nascimento', '').strip() or None
        notas = request.POST.get('notas', '').strip()

        if not nome:
            erros['nome'] = 'Nome obrigatório.'
        if not bi:
            erros['bi'] = 'Nº BI obrigatório.'
        else:
            bi_limpo = limpar_bi(bi)
            if not re.fullmatch(BI_REGEX, bi_limpo):
                erros['bi'] = 'BI inválido. Formato: 9 dígitos + 2 letras + 3 dígitos (ex: 003456789LA045).'
            else:
                qs_bi = Funcionario.objects.filter(bi=bi_limpo)
                if instancia:
                    qs_bi = qs_bi.exclude(pk=instancia.pk)
                if qs_bi.exists():
                    erros['bi'] = 'Já existe um funcionário com este BI.'
                else:
                    bi = bi_limpo
        if not telefone:
            erros['telefone'] = 'Telefone obrigatório.'
        else:
            tel_limpo = limpar_telefone(telefone)
            if not re.fullmatch(TELEFONE_REGEX, tel_limpo):
                erros['telefone'] = 'Telefone inválido. Formato: 9XX XXX XXX (ex: 923 456 789).'
        if not data_admissao:
            erros['data_admissao'] = 'Data de admissão obrigatória.'
        import re as _re
        if nib_raw and not _re.fullmatch(r'\d{21}', nib_raw):
            erros['nib'] = 'NIB inválido. Deve ter exactamente 21 dígitos.'
        if iban_raw and not _re.fullmatch(r'AO\d{23}', iban_raw):
            erros['iban'] = 'IBAN inválido. Formato: AO + 23 dígitos (total 25 caracteres).'

        if not erros:
            salario = salario_raw if salario_raw else None
            utilizador = None
            if utilizador_id:
                try:
                    utilizador = User.objects.get(pk=utilizador_id)
                except User.DoesNotExist:
                    logger.warning('Utilizador ID %s não encontrado ao criar/editar funcionário.', utilizador_id)
                    erros['utilizador'] = 'Utilizador seleccionado não existe.'
            if erros:
                post = request.POST
                return render(request, 'gestao/funcionario_form.html', {
                    'active_page': 'funcionarios',
                    'utilizadores': utilizadores_disponiveis,
                    'cargo_choices': Funcionario.CARGO_CHOICES,
                    'turno_choices': Funcionario.TURNO_CHOICES,
                    'erros': erros,
                    'post': post,
                    'instancia': instancia,
                })

            try:
                percentagem_comissao = float(percentagem_comissao_raw.replace(',', '.'))
                if not (0 <= percentagem_comissao <= 100):
                    percentagem_comissao = 2.0
            except (ValueError, TypeError):
                percentagem_comissao = 2.0
            if cargo not in Funcionario.CARGO_COM_COMISSAO:
                percentagem_comissao = 0.0
            try:
                subsidio_alimentacao = float(subsidio_alimentacao_raw.replace(',', '.'))
                subsidio_alimentacao = max(0, subsidio_alimentacao)
            except (ValueError, TypeError):
                subsidio_alimentacao = 0.0
            try:
                subsidio_transporte = float(subsidio_transporte_raw.replace(',', '.'))
                subsidio_transporte = max(0, subsidio_transporte)
            except (ValueError, TypeError):
                subsidio_transporte = 0.0

            if instancia:
                instancia.nome = nome
                instancia.bi = bi
                instancia.telefone = telefone
                instancia.email = email
                instancia.cargo = cargo
                instancia.turno = turno
                instancia.salario = salario
                instancia.percentagem_comissao = percentagem_comissao
                instancia.subsidio_alimentacao = subsidio_alimentacao
                instancia.subsidio_transporte = subsidio_transporte
                instancia.banco = banco
                instancia.nib = nib_raw
                instancia.iban = iban_raw
                instancia.data_nascimento = data_nascimento
                instancia.data_admissao = data_admissao
                instancia.activo = activo
                instancia.utilizador = utilizador
                instancia.notas = notas
                if request.POST.get('foto_remover') == '1':
                    instancia.foto.delete(save=False)
                    instancia.foto = None
                if request.FILES.get('foto'):
                    instancia.foto = request.FILES['foto']
                instancia.save()
                messages.success(request, f'Funcionário "{nome}" actualizado.')
            else:
                Funcionario.objects.create(
                    nome=nome, bi=bi, telefone=telefone, email=email,
                    cargo=cargo, turno=turno, salario=salario,
                    percentagem_comissao=percentagem_comissao,
                    subsidio_alimentacao=subsidio_alimentacao,
                    subsidio_transporte=subsidio_transporte,
                    banco=banco, nib=nib_raw, iban=iban_raw,
                    data_nascimento=data_nascimento,
                    data_admissao=data_admissao, activo=activo,
                    utilizador=utilizador, notas=notas,
                    foto=request.FILES.get('foto') or None,
                )
                messages.success(request, f'Funcionário "{nome}" registado com sucesso.')
            return redirect('gestao_funcionarios')

        # Reenviar valores preenchidos
        post = request.POST
        return render(request, 'gestao/funcionario_form.html', {
            'active_page': 'funcionarios',
            'utilizadores': utilizadores_disponiveis,
            'cargo_choices': Funcionario.CARGO_CHOICES,
            'turno_choices': Funcionario.TURNO_CHOICES,
            'banco_choices': Funcionario.BANCO_CHOICES,
            'erros': erros,
            'post': post,
            'instancia': instancia,
        })

    initial = {}
    if instancia:
        initial = {
            'nome': instancia.nome, 'bi': instancia.bi, 'telefone': instancia.telefone,
            'email': instancia.email, 'cargo': instancia.cargo, 'turno': instancia.turno,
            'salario': instancia.salario or '', 'data_admissao': instancia.data_admissao.isoformat() if instancia.data_admissao else '',
            'data_nascimento': instancia.data_nascimento.isoformat() if instancia.data_nascimento else '',
            'activo': 'on' if instancia.activo else '',
            'utilizador': str(instancia.utilizador_id) if instancia.utilizador_id else '',
            'notas': instancia.notas,
        }

    return render(request, 'gestao/funcionario_form.html', {
        'active_page': 'funcionarios',
        'utilizadores': utilizadores_disponiveis,
        'cargo_choices': Funcionario.CARGO_CHOICES,
        'turno_choices': Funcionario.TURNO_CHOICES,
        'banco_choices': Funcionario.BANCO_CHOICES,
        'erros': erros,
        'post': initial,
        'instancia': instancia,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_funcionario_apagar(request, pk):
    f = get_object_or_404(Funcionario, pk=pk)
    if request.method == 'POST':
        nome = f.nome
        f.delete()
        messages.success(request, f'Funcionário "{nome}" eliminado.')
    return redirect('gestao_funcionarios')


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_funcionario_criar_utilizador(request, pk):
    """Cria uma conta de utilizador (User) e associa ao funcionário.
    Apenas administradores (superuser) podem usar esta acção.
    """
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem criar utilizadores.')
        return redirect('gestao_funcionario_detalhe', pk=pk)

    funcionario = get_object_or_404(Funcionario, pk=pk)

    if funcionario.utilizador_id:
        messages.warning(request, 'Este funcionário já tem um utilizador associado.')
        return redirect('gestao_funcionario_detalhe', pk=pk)

    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''
        password2 = request.POST.get('password2') or ''
        nivel = (request.POST.get('nivel') or 'gerente').strip().lower()
        if nivel not in ('administrador', 'gerente', 'operador'):
            nivel = 'gerente'

        erros = []
        if not username:
            erros.append('Nome de utilizador é obrigatório.')
        elif User.objects.filter(username__iexact=username).exists():
            erros.append('Já existe um utilizador com esse nome.')
        if len(password) < 6:
            erros.append('A senha deve ter pelo menos 6 caracteres.')
        if password != password2:
            erros.append('As senhas não coincidem.')

        if erros:
            for e in erros:
                messages.error(request, e)
        else:
            user = User.objects.create_user(
                username=username,
                password=password,
                email=funcionario.email or '',
                first_name=funcionario.nome.split(' ')[0] if funcionario.nome else '',
            )
            if nivel == 'administrador':
                user.is_staff = True
                user.is_superuser = True
            elif nivel == 'gerente':
                user.is_staff = True
                user.is_superuser = False
            else:  # operador
                user.is_staff = True
                user.is_superuser = False
            user.save()
            if nivel == 'operador':
                grupo_op, _ = Group.objects.get_or_create(name='Operador')
                user.groups.add(grupo_op)
            funcionario.utilizador = user
            funcionario.save(update_fields=['utilizador'])

            # Mensagem para envio (email + SMS)
            primeiro_nome = (funcionario.nome.split(' ')[0] if funcionario.nome else username)
            mensagem = (
                f"Olá {primeiro_nome}, os seus dados de acesso para a Décent Privé são:\n"
                f"Utilizador: {username}\n"
                f"Senha: {password}\n"
                f"Nível: {nivel.capitalize()}\n"
                f"Aceda em: {request.build_absolute_uri('/gestao/login/')}"
            )

            # --- ENVIO POR EMAIL ---
            email_status = 'nao_enviado'
            email_erro = ''
            if funcionario.email:
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings as dj_settings
                    send_mail(
                        subject='Décent Privé — Dados de acesso',
                        message=mensagem,
                        from_email=getattr(dj_settings, 'DEFAULT_FROM_EMAIL', None),
                        recipient_list=[funcionario.email],
                        fail_silently=False,
                    )
                    email_status = 'enviado'
                except Exception as exc:
                    email_status = 'erro'
                    email_erro = str(exc)
                    logger.exception('Falha ao enviar email de credenciais a %s: %s', funcionario.email, exc)

            # --- ENVIO POR SMS ---
            sms_status = 'nao_enviado'
            sms_erro = ''
            if funcionario.telefone:
                try:
                    from django.conf import settings as dj_settings
                    gateway_url = getattr(dj_settings, 'SMS_GATEWAY_URL', '') or ''
                    if gateway_url:
                        import requests  # noqa: WPS433
                        resp = requests.post(
                            gateway_url,
                            data={
                                'to': funcionario.telefone,
                                'message': mensagem,
                                'sender': getattr(dj_settings, 'SMS_SENDER', 'DécentPrive'),
                                'api_key': getattr(dj_settings, 'SMS_GATEWAY_API_KEY', ''),
                            },
                            timeout=10,
                        )
                        if 200 <= resp.status_code < 300:
                            sms_status = 'enviado'
                        else:
                            sms_status = 'erro'
                            sms_erro = f'HTTP {resp.status_code}'
                    else:
                        # Sem gateway configurado: regista no log do servidor
                        import logging
                        logging.getLogger(__name__).info(
                            'SMS (simulado) para %s:\n%s', funcionario.telefone, mensagem,
                        )
                        sms_status = 'simulado'
                except Exception as exc:
                    sms_status = 'erro'
                    sms_erro = str(exc)
                    logger.exception('Falha ao enviar SMS de credenciais a %s: %s', funcionario.telefone, exc)

            messages.success(
                request,
                f'Utilizador "{username}" ({nivel}) criado e associado a {funcionario.nome}.',
            )
            return render(request, 'gestao/funcionario_utilizador_sucesso.html', {
                'funcionario': funcionario,
                'username': username,
                'password': password,
                'nivel': nivel,
                'mensagem': mensagem,
                'email_destino': funcionario.email,
                'email_status': email_status,
                'email_erro': email_erro,
                'telefone_destino': funcionario.telefone,
                'sms_status': sms_status,
                'sms_erro': sms_erro,
                'active_page': 'funcionarios',
            })

    return render(request, 'gestao/funcionario_criar_utilizador.html', {
        'funcionario': funcionario,
        'active_page': 'funcionarios',
        'post': request.POST if request.method == 'POST' else {},
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_funcionario_detalhe(request, pk):
    funcionario = get_object_or_404(Funcionario, pk=pk)

    vendas_qs = Encomenda.objects.none()
    if funcionario.utilizador_id:
        vendas_qs = Encomenda.objects.filter(vendido_por_id=funcionario.utilizador_id)

    total_vendas = vendas_qs.count()
    total_finalizadas = vendas_qs.filter(status='finalizada').count()
    total_em_curso = vendas_qs.filter(status='em_curso').count()
    total_canceladas = vendas_qs.filter(status='cancelada').count()

    total_facturado = 0
    finalizadas_qs = vendas_qs.filter(status='finalizada').prefetch_related('itens')
    for e in finalizadas_qs:
        total_facturado += sum((i.preco_unitario or 0) * i.quantidade for i in e.itens.all())
    ticket_medio = (total_facturado / total_finalizadas) if total_finalizadas else 0

    encomendas = list(vendas_qs.order_by('-criado_em')[:10].prefetch_related('itens'))
    for e in encomendas:
        e.total = sum((i.preco_unitario or 0) * i.quantidade for i in e.itens.all())

    # Tempo de casa
    anos_de_casa = None
    if funcionario.data_admissao:
        hoje = timezone.now().date()
        delta = hoje - funcionario.data_admissao
        anos_de_casa = round(delta.days / 365.25, 1)

    return render(request, 'gestao/funcionario_detalhe.html', {
        'active_page': 'funcionarios',
        'funcionario': funcionario,
        'total_vendas': total_vendas,
        'total_finalizadas': total_finalizadas,
        'total_em_curso': total_em_curso,
        'total_canceladas': total_canceladas,
        'total_facturado': total_facturado,
        'ticket_medio': ticket_medio,
        'encomendas': encomendas,
        'anos_de_casa': anos_de_casa,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_produto_criar(request, pk=None):
    instancia = get_object_or_404(Produto, pk=pk) if pk else None
    # Garantir que as duas categorias fixas existem e obter os IDs
    cat_ids = []
    for nome_cat in ['Nicho', 'Diversos']:
        try:
            obj = Categoria.objects.filter(nome__iexact=nome_cat).first()
            if not obj:
                obj = Categoria.objects.create(
                    nome=nome_cat,
                    slug=gerar_slug_unico(Categoria, nome_cat),
                )
            cat_ids.append(obj.id)
        except Exception as e:
            logger.error('Erro ao garantir categoria %s: %s', nome_cat, e)
    categorias = Categoria.objects.filter(id__in=cat_ids).order_by('nome')
    logger.warning('[PRODUTO_FORM] cat_ids=%s, categorias_count=%d', cat_ids, categorias.count())
    genero_choices = Produto.GENERO_CHOICES
    concentracao_choices = Produto.CONCENTRACAO_CHOICES
    origem_choices = Produto.ORIGEM_CHOICES
    erros = {}
    post = {}

    if request.method == 'POST':
        post = request.POST
        nome = post.get('nome', '').strip()
        marca = post.get('marca', '').strip()
        descricao = post.get('descricao', '').strip()
        preco_venda_raw = post.get('preco_venda', '').strip() or post.get('preco', '').strip()
        preco_compra_raw = post.get('preco_compra', '0').strip() or '0'
        quantidade = post.get('quantidade', '').strip()
        categoria_id = post.get('categoria', '').strip()
        genero = post.get('genero', 'unissex')
        concentracao = post.get('concentracao', '').strip()
        origem = post.get('origem', 'europeu')
        stock_raw = post.get('stock', '0').strip()
        disponivel = post.get('disponivel') == 'on'
        codigo_barras = post.get('codigo_barras', '').strip() or None
        essencia = post.get('essencia', '').strip()
        desconto_raw = post.get('desconto_percentagem', '0').strip() or '0'

        if not nome:
            erros['nome'] = 'Nome obrigatório.'
        if not marca:
            erros['marca'] = 'Marca obrigatória.'
        if not preco_venda_raw:
            erros['preco_venda'] = 'Preço de venda obrigatório.'
        else:
            try:
                preco_venda = float(preco_venda_raw.replace(',', '.'))
                if preco_venda < 0:
                    erros['preco_venda'] = 'Preço de venda deve ser positivo.'
            except ValueError:
                erros['preco_venda'] = 'Preço de venda inválido.'
        try:
            preco_compra = float(preco_compra_raw.replace(',', '.'))
            if preco_compra < 0:
                erros['preco_compra'] = 'Preço de compra deve ser positivo.'
        except ValueError:
            erros['preco_compra'] = 'Preço de compra inválido.'
            preco_compra = 0
        try:
            desconto_percentagem = int(desconto_raw)
            if desconto_percentagem < 0 or desconto_percentagem > 100:
                erros['desconto_percentagem'] = 'Desconto deve estar entre 0 e 100.'
            elif desconto_percentagem > 0 and not erros.get('preco_venda'):
                preco_promocional = round(preco_venda * (100 - desconto_percentagem) / 100, 2)
                if preco_promocional >= preco_venda:
                    erros['desconto_percentagem'] = 'O preço promocional deve ser inferior ao preço original.'
        except ValueError:
            erros['desconto_percentagem'] = 'Percentagem de desconto inválida.'
            desconto_percentagem = 0
        if codigo_barras:
            qs_cb = Produto.objects.filter(codigo_barras=codigo_barras)
            if instancia:
                qs_cb = qs_cb.exclude(pk=instancia.pk)
            if qs_cb.exists():
                erros['codigo_barras'] = 'Já existe um produto com este código de barras.'

        if not erros:
            try:
                stock = int(stock_raw) if stock_raw else 0
                if stock < 0:
                    erros['stock'] = 'Stock não pode ser negativo.'
            except ValueError:
                erros['stock'] = 'Valor de stock inválido.'
                stock = 0

            categoria = None
            if categoria_id:
                try:
                    categoria = Categoria.objects.get(pk=categoria_id)
                except Categoria.DoesNotExist:
                    logger.warning('Categoria ID %s não encontrada ao criar/editar produto.', categoria_id)
                    erros['categoria'] = 'Categoria seleccionada não existe.'
            if erros:
                return render(request, 'gestao/produto_form.html', {
                    'active_page': 'produtos',
                    'categorias': categorias,
                    'genero_choices': genero_choices,
                    'concentracao_choices': concentracao_choices,
                    'origem_choices': origem_choices,
                    'erros': erros,
                    'post': post,
                    'instancia': instancia,
                })

            imagem = request.FILES.get('imagem') or None

            if instancia:
                instancia.nome = nome
                instancia.marca = marca
                instancia.descricao = descricao
                instancia.preco_venda = preco_venda
                instancia.preco_compra = preco_compra
                instancia.quantidade = quantidade
                instancia.categoria = categoria
                instancia.genero = genero
                instancia.concentracao = concentracao
                instancia.origem = origem
                instancia.stock = stock
                instancia.disponivel = disponivel
                if imagem:
                    instancia.imagem = imagem
                instancia.codigo_barras = codigo_barras
                instancia.essencia = essencia
                instancia.desconto_percentagem = desconto_percentagem
                instancia.save()
                messages.success(request, f'Produto "{nome}" actualizado.')
            else:
                Produto.objects.create(
                    nome=nome,
                    marca=marca,
                    descricao=descricao,
                    preco_venda=preco_venda,
                    preco_compra=preco_compra,
                    quantidade=quantidade,
                    categoria=categoria,
                    genero=genero,
                    concentracao=concentracao,
                    origem=origem,
                    stock=stock,
                    disponivel=disponivel,
                    imagem=imagem,
                    codigo_barras=codigo_barras,
                    essencia=essencia,
                    desconto_percentagem=desconto_percentagem,
                )
                messages.success(request, f'Produto "{nome}" criado com sucesso.')
            return redirect('gestao_produtos')

    if instancia and not post:
        post = {
            'nome': instancia.nome, 'marca': instancia.marca, 'descricao': instancia.descricao,
            'preco_venda': str(instancia.preco_venda), 'preco_compra': str(instancia.preco_compra),
            'quantidade': instancia.quantidade,
            'categoria': str(instancia.categoria_id) if instancia.categoria_id else '',
            'genero': instancia.genero,
            'concentracao': instancia.concentracao,
            'origem': instancia.origem,
            'stock': str(instancia.stock),
            'disponivel': 'on' if instancia.disponivel else '',
            'codigo_barras': instancia.codigo_barras or '',
            'essencia': instancia.essencia,
            'desconto_percentagem': str(instancia.desconto_percentagem),
        }

    return render(request, 'gestao/produto_form.html', {
        'active_page': 'produtos',
        'categorias': categorias,
        'genero_choices': genero_choices,
        'concentracao_choices': concentracao_choices,
        'origem_choices': origem_choices,
        'erros': erros,
        'post': post,
        'instancia': instancia,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_produto_apagar(request, pk):
    p = get_object_or_404(Produto, pk=pk)
    if request.method == 'POST':
        nome = p.nome
        p.delete()
        messages.success(request, f'Produto "{nome}" eliminado.')
    return redirect('gestao_produtos')


@staff_member_required(login_url='/gestao/login/')
def gestao_produto_detalhe(request, pk):
    produto = get_object_or_404(Produto, pk=pk)
    movimentos = MovimentoStock.objects.filter(produto=produto).order_by('-criado_em')[:50]
    return render(request, 'gestao/produto_detalhe.html', {
        'active_page': 'produtos',
        'produto': produto,
        'movimentos': movimentos,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_produto_entrada_stock(request, pk):
    """Adicionar entrada manual de stock (recebimento de fornecedor)."""
    produto = get_object_or_404(Produto, pk=pk)
    erros = {}
    
    if request.method == 'POST':
        quantidade_raw = request.POST.get('quantidade', '').strip()
        descricao = request.POST.get('descricao', '').strip()
        
        if not quantidade_raw:
            erros['quantidade'] = 'Quantidade obrigatória.'
        else:
            try:
                quantidade = int(quantidade_raw)
                if quantidade <= 0:
                    erros['quantidade'] = 'A quantidade deve ser positiva.'
            except ValueError:
                erros['quantidade'] = 'Quantidade inválida.'
        
        if not descricao:
            erros['descricao'] = 'Descrição obrigatória (ex: "Recebimento fornecedor" ou "Ajuste manual").'
        
        if not erros:
            # Actualizar stock do produto
            produto.stock += quantidade
            produto.save(update_fields=['stock'])
            
            # Registar movimento
            MovimentoStock.objects.create(
                produto=produto,
                tipo='entrada',
                quantidade=quantidade,
                descricao=descricao,
                criado_por=request.user if request.user.is_authenticated else None,
            )
            
            messages.success(
                request, 
                f'Stock de "{produto.nome}" aumentado em {quantidade} unidades. '
                f'Total agora: {produto.stock}'
            )
            return redirect('gestao_produto_detalhe', pk=pk)
    
    movimentos = MovimentoStock.objects.filter(produto=produto).order_by('-criado_em')[:20]
    
    return render(request, 'gestao/produto_entrada_stock.html', {
        'active_page': 'produtos',
        'produto': produto,
        'movimentos': movimentos,
        'erros': erros,
        'post': request.POST if request.method == 'POST' else {},
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_encomendas(request):
    """Pedidos online (encomendas em curso, separadas das vendas de balcão)."""
    if request.method == 'POST':
        enc_id = request.POST.get('encomenda_id')
        novo_status = request.POST.get('status')
        status_validos = {s[0] for s in Encomenda.STATUS_CHOICES}
        if enc_id and novo_status:
            if novo_status not in status_validos:
                messages.error(request, f'Estado "{novo_status}" inválido.')
                return redirect('gestao_encomendas')
            enc = get_object_or_404(Encomenda, pk=enc_id)
            enc.status = novo_status
            enc.save(update_fields=['status', 'atualizado_em'])
            if novo_status == 'finalizada':
                _registar_venda_no_caixa(enc, request.user)
            messages.success(request, f'Encomenda #{enc.pk} actualizada para "{enc.get_status_display()}".')
        return redirect('gestao_encomendas')

    # Pedidos em curso (não finalizados nem cancelados)
    abertas_status = ['em_curso']
    encomendas = (Encomenda.objects
                  .filter(status__in=abertas_status)
                  .prefetch_related('itens')
                  .order_by('-criado_em'))

    contadores = {s: 0 for s in abertas_status}
    for e in encomendas:
        contadores[e.status] = contadores.get(e.status, 0) + 1

    return render(request, 'gestao/encomendas.html', {
        'active_page': 'encomendas',
        'encomendas': encomendas,
        'status_choices': Encomenda.STATUS_CHOICES,
        'contadores': contadores,
        'total_abertas': len(encomendas),
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_caixa(request):
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    from datetime import timedelta
    inicio_semana = hoje - timedelta(days=hoje.weekday())

    enc_hoje = Encomenda.objects.filter(criado_em__date=hoje, status='finalizada').prefetch_related('itens')
    enc_semana = Encomenda.objects.filter(criado_em__date__gte=inicio_semana, status='finalizada').prefetch_related('itens')
    enc_mes = Encomenda.objects.filter(criado_em__date__gte=inicio_mes, status='finalizada').prefetch_related('itens')

    def total_enc(qs):
        return sum(sum(i.subtotal() for i in e.itens.all()) for e in qs)

    # Top produtos do mês
    top_produtos = (
        ItemEncomenda.objects
        .filter(encomenda__criado_em__date__gte=inicio_mes, encomenda__status='finalizada')
        .values('nome_produto')
        .annotate(total_qty=Sum('quantidade'), total_valor=Sum('preco_unitario'))
        .order_by('-total_qty')[:8]
    )

    # Vendas finalizadas hoje
    ultimas_vendas = (
        Encomenda.objects
        .filter(status='finalizada', criado_em__date=hoje)
        .prefetch_related('itens')
        .order_by('-criado_em')
    )
    for v in ultimas_vendas:
        v.total_calc = sum(i.subtotal() for i in v.itens.all())

    return render(request, 'gestao/caixa.html', {
        'active_page': 'caixa',
        'hoje': hoje,
        'receita_hoje': total_enc(enc_hoje),
        'vendas_hoje': enc_hoje.count(),
        'receita_semana': total_enc(enc_semana),
        'vendas_semana': enc_semana.count(),
        'receita_mes': total_enc(enc_mes),
        'vendas_mes': enc_mes.count(),
        'top_produtos': top_produtos,
        'ultimas_vendas': ultimas_vendas,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_clientes(request):
    clientes = Cliente.objects.order_by('-pontos', 'nome')
    return render(request, 'gestao/clientes.html', {
        'active_page': 'clientes',
        'clientes': clientes,
    })


@staff_member_required(login_url='/gestao/login/')
def gestao_cliente_criar(request, pk=None):
    instancia = get_object_or_404(Cliente, pk=pk) if pk else None
    erros = {}
    post = {}
    if request.method == 'POST':
        post = request.POST
        nome = post.get('nome', '').strip()
        telefone = post.get('telefone', '').strip()
        email = post.get('email', '').strip()
        notas = post.get('notas', '').strip()
        data_nasc_raw = post.get('data_nascimento', '').strip()
        data_nascimento = None
        if data_nasc_raw:
            from datetime import datetime as _dt
            try:
                data_nascimento = _dt.strptime(data_nasc_raw, '%Y-%m-%d').date()
            except ValueError:
                erros['data_nascimento'] = 'Data de nascimento inválida.'

        if not nome:
            erros['nome'] = 'Nome obrigatório.'
        if not telefone:
            erros['telefone'] = 'Telefone obrigatório.'
        else:
            qs_tel = Cliente.objects.filter(telefone=telefone)
            if instancia:
                qs_tel = qs_tel.exclude(pk=instancia.pk)
            if qs_tel.exists():
                erros['telefone'] = 'Já existe um cliente com este telefone.'

        if not erros:
            if instancia:
                instancia.nome = nome
                instancia.telefone = telefone
                instancia.email = email
                instancia.notas = notas
                instancia.data_nascimento = data_nascimento
                instancia.save()
                messages.success(request, f'Cliente "{nome}" actualizado.')
            else:
                Cliente.objects.create(nome=nome, telefone=telefone, email=email,
                                        notas=notas, data_nascimento=data_nascimento)
                messages.success(request, f'Cliente "{nome}" registado com sucesso.')
            return redirect('gestao_clientes')

    if instancia and not post:
        post = {
            'nome': instancia.nome, 'telefone': instancia.telefone,
            'email': instancia.email, 'notas': instancia.notas,
            'data_nascimento': instancia.data_nascimento.strftime('%Y-%m-%d') if instancia.data_nascimento else '',
        }

    return render(request, 'gestao/cliente_form.html', {
        'active_page': 'clientes',
        'erros': erros,
        'post': post,
        'instancia': instancia,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_cliente_apagar(request, pk):
    c = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        nome = c.nome
        c.delete()
        messages.success(request, f'Cliente "{nome}" eliminado.')
    return redirect('gestao_clientes')


@staff_member_required(login_url='/gestao/login/')
def gestao_cliente_detalhe(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    encomendas_qs = Encomenda.objects.filter(telefone=cliente.telefone).order_by('-criado_em')
    encomendas = list(encomendas_qs[:20])

    total_encomendas = encomendas_qs.count()
    total_finalizadas = encomendas_qs.filter(status='finalizada').count()
    total_em_curso = encomendas_qs.filter(status='em_curso').count()
    total_canceladas = encomendas_qs.filter(status='cancelada').count()

    total_gasto = 0
    for e in encomendas_qs.filter(status='finalizada'):
        total_gasto += float(e.total() or 0)

    ticket_medio = (total_gasto / total_finalizadas) if total_finalizadas else 0
    ultima = encomendas[0] if encomendas else None

    return render(request, 'gestao/cliente_detalhe.html', {
        'active_page': 'clientes',
        'cliente': cliente,
        'encomendas': encomendas,
        'total_encomendas': total_encomendas,
        'total_finalizadas': total_finalizadas,
        'total_em_curso': total_em_curso,
        'total_canceladas': total_canceladas,
        'total_gasto': total_gasto,
        'ticket_medio': ticket_medio,
        'ultima_encomenda': ultima,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_categorias(request):
    erros = {}
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if not nome:
            erros['nome'] = 'Nome obrigatório.'
        else:
            slug = gerar_slug_unico(Categoria, nome)
            Categoria.objects.create(nome=nome, slug=slug)
            messages.success(request, f'Categoria "{nome}" criada.')
            return redirect('gestao_categorias')

    categorias = Categoria.objects.annotate(num_produtos=Count('produtos')).order_by('nome')
    return render(request, 'gestao/categorias.html', {
        'active_page': 'categorias',
        'categorias': categorias,
        'erros': erros,
        'post': request.POST if request.method == 'POST' else {},
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_categoria_apagar(request, pk):
    cat = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        nome = cat.nome
        cat.delete()
        messages.success(request, f'Categoria "{nome}" eliminada.')
    return redirect('gestao_categorias')


@staff_member_required(login_url='/gestao/login/')
def gestao_api_barcode(request):
    codigo = request.GET.get('codigo', '').strip()
    if not codigo:
        return JsonResponse({'encontrado': False, 'erro': 'Código em falta.'})
    try:
        p = Produto.objects.get(codigo_barras=codigo)
        return JsonResponse({
            'encontrado': True,
            'id': p.pk,
            'nome': p.nome,
            'marca': p.marca,
            'preco': float(p.preco),
            'stock': p.stock,
            'disponivel': p.disponivel,
        })
    except Produto.DoesNotExist:
        return JsonResponse({'encontrado': False, 'erro': 'Produto não encontrado para este código.'})


@staff_member_required(login_url='/gestao/login/')
def gestao_api_session_ping(request):
    """
    Endpoint para renovar a sessão quando o utilizador está ativo.
    Chamado periodicamente pelo JavaScript de monitorização de atividade.
    """
    if request.method == 'POST':
        # Renova a sessão modificando a data de atualização
        request.session.modified = True
        return JsonResponse({'status': 'ok', 'timestamp': timezone.now().isoformat()})
    return JsonResponse({'status': 'error'}, status=405)


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_relatorio(request):
    """Relatório consolidado de vendas e produtos."""
    from datetime import timedelta
    from django.db.models import F, FloatField, ExpressionWrapper

    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    ha_30_dias = hoje - timedelta(days=30)

    # Período seleccionável (querystring ?periodo=hoje|semana|mes|todos)
    periodo = request.GET.get('periodo', 'mes')
    if periodo == 'hoje':
        data_inicio = hoje
        rotulo = 'Hoje'
    elif periodo == 'semana':
        data_inicio = inicio_semana
        rotulo = 'Esta Semana'
    elif periodo == 'todos':
        data_inicio = None
        rotulo = 'Todo o Período'
    else:
        periodo = 'mes'
        data_inicio = inicio_mes
        rotulo = 'Este Mês'

    vendas_qs = Encomenda.objects.filter(status='finalizada')
    if data_inicio:
        vendas_qs = vendas_qs.filter(criado_em__date__gte=data_inicio)

    total_vendas = vendas_qs.count()
    receita_total = sum(
        sum(i.subtotal() for i in e.itens.all())
        for e in vendas_qs.prefetch_related('itens')
    )
    ticket_medio = (receita_total / total_vendas) if total_vendas else 0

    # Encomendas online em curso (independentes do período)
    pendentes_online = Encomenda.objects.filter(
        status='em_curso', vendido_por__isnull=True
    ).count()
    canceladas = Encomenda.objects.filter(
        status='cancelada',
        **({'criado_em__date__gte': data_inicio} if data_inicio else {})
    ).count()

    # Top 10 produtos mais vendidos
    itens_qs = ItemEncomenda.objects.filter(encomenda__status='finalizada')
    if data_inicio:
        itens_qs = itens_qs.filter(encomenda__criado_em__date__gte=data_inicio)

    top_produtos = (
        itens_qs.values('nome_produto')
        .annotate(
            qty=Sum('quantidade'),
            receita=Sum(
                ExpressionWrapper(F('preco_unitario') * F('quantidade'),
                                  output_field=FloatField())
            ),
        )
        .order_by('-qty')[:10]
    )

    # Vendedores
    top_vendedores = (
        vendas_qs.exclude(vendido_por__isnull=True)
        .values('vendido_por__username', 'vendido_por__first_name', 'vendido_por__last_name')
        .annotate(num_vendas=Count('id'))
        .order_by('-num_vendas')[:10]
    )
    for v in top_vendedores:
        nome = f"{v['vendido_por__first_name']} {v['vendido_por__last_name']}".strip()
        v['nome'] = nome or v['vendido_por__username']

    # Stock crítico (≤ 5 unidades)
    stock_critico = Produto.objects.filter(stock__lte=5).order_by('stock', 'nome')[:15]
    total_produtos = Produto.objects.count()
    produtos_indisponiveis = Produto.objects.filter(disponivel=False).count()
    valor_stock = sum(float(p.preco_venda) * p.stock for p in Produto.objects.all())

    # Origem das vendas
    vendas_balcao = vendas_qs.exclude(vendido_por__isnull=True).count()
    vendas_online = vendas_qs.filter(vendido_por__isnull=True).count()

    return render(request, 'gestao/relatorio.html', {
        'active_page': 'relatorio',
        'periodo': periodo,
        'rotulo': rotulo,
        'total_vendas': total_vendas,
        'receita_total': receita_total,
        'ticket_medio': ticket_medio,
        'pendentes_online': pendentes_online,
        'canceladas': canceladas,
        'top_produtos': top_produtos,
        'top_vendedores': top_vendedores,
        'stock_critico': stock_critico,
        'total_produtos': total_produtos,
        'produtos_indisponiveis': produtos_indisponiveis,
        'valor_stock': valor_stock,
        'vendas_balcao': vendas_balcao,
        'vendas_online': vendas_online,
    })


# ---------------------------------------------------------------------------
# Financeiro — Movimentos de Caixa
# ---------------------------------------------------------------------------

def _financeiro_listar(request, tipo):
    """Helper genérico para Entradas/Saídas de caixa."""
    erros = {}
    categoria_choices = [
        c for c in MovimentoCaixa.CATEGORIA_CHOICES
        if (tipo == 'entrada' and c[0] in ('reforco', 'investimento', 'outro_ganho'))
        or (tipo == 'saida' and c[0] in ('despesa', 'salario', 'renda', 'utilidades',
                                          'fornecedor', 'imposto', 'manutencao', 'outro_custo'))
    ]

    if request.method == 'POST':
        descricao = request.POST.get('descricao', '').strip()
        categoria = request.POST.get('categoria', '').strip()
        valor_raw = request.POST.get('valor', '').strip().replace(',', '.')
        notas = request.POST.get('notas', '').strip()

        try:
            valor = float(valor_raw)
        except (ValueError, TypeError):
            valor = 0
            erros['valor'] = 'Valor inválido. Introduza um número.'

        if not descricao:
            erros['descricao'] = 'Descrição obrigatória.'
        if categoria not in dict(categoria_choices):
            erros['categoria'] = 'Seleccione uma categoria válida.'
        if valor <= 0:
            erros['valor'] = 'Valor deve ser maior que zero.'

        if not erros:
            MovimentoCaixa.objects.create(
                tipo=tipo,
                categoria=categoria,
                valor=valor,
                descricao=descricao,
                notas=notas,
                criado_por=request.user if request.user.is_authenticated else None,
            )
            messages.success(request, f'{"Entrada" if tipo == "entrada" else "Saída"} de caixa registada.')
            return redirect('gestao_caixa_entrada' if tipo == 'entrada' else 'gestao_caixa_saida')

    movimentos = MovimentoCaixa.objects.filter(tipo=tipo).exclude(categoria='venda').select_related('criado_por').order_by('-data')[:100]
    total = sum(float(m.valor) for m in movimentos)

    # Saldo geral do caixa (todas as entradas − todas as saídas)
    todas_entradas = sum(float(m.valor) for m in MovimentoCaixa.objects.filter(tipo='entrada'))
    todas_saidas = sum(float(m.valor) for m in MovimentoCaixa.objects.filter(tipo='saida'))

    return render(request, f'gestao/financeiro_{tipo}.html', {
        'active_page': f'financeiro_{tipo}',
        'movimentos': movimentos,
        'total': total,
        'categoria_choices': categoria_choices,
        'erros': erros,
        'post': request.POST if request.method == 'POST' else {},
        'saldo_geral': todas_entradas - todas_saidas,
        'todas_entradas': todas_entradas,
        'todas_saidas': todas_saidas,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_caixa_entrada(request):
    return _financeiro_listar(request, 'entrada')


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_caixa_saida(request):
    return _financeiro_listar(request, 'saida')


# ---------------------------------------------------------------------------
# Lucro — Análise de margem (preço de venda − preço de compra)
# ---------------------------------------------------------------------------

@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_lucro(request):
    periodo = request.GET.get('periodo', 'mes')
    hoje = timezone.localdate()
    if periodo == 'hoje':
        inicio = hoje
    elif periodo == 'semana':
        inicio = hoje - timezone.timedelta(days=7)
    elif periodo == 'todos':
        inicio = None
    else:
        periodo = 'mes'
        inicio = hoje.replace(day=1)

    qs = ItemEncomenda.objects.filter(encomenda__status='finalizada').select_related('encomenda', 'produto')
    if inicio:
        qs = qs.filter(encomenda__criado_em__date__gte=inicio)

    receita_total = 0.0
    custo_total = 0.0
    qtd_total = 0
    por_produto = {}

    for item in qs:
        preco = float(item.preco_unitario or 0)
        custo = float(item.preco_custo_unitario or 0)
        qty = item.quantidade or 0
        sub_receita = preco * qty
        sub_custo = custo * qty
        receita_total += sub_receita
        custo_total += sub_custo
        qtd_total += qty

        chave = item.produto_id or item.nome_produto
        if chave not in por_produto:
            por_produto[chave] = {
                'nome': item.nome_produto,
                'marca': item.produto.marca if item.produto else '',
                'qty': 0,
                'receita': 0.0,
                'custo': 0.0,
            }
        por_produto[chave]['qty'] += qty
        por_produto[chave]['receita'] += sub_receita
        por_produto[chave]['custo'] += sub_custo

    for d in por_produto.values():
        d['lucro'] = d['receita'] - d['custo']
        d['margem'] = (d['lucro'] / d['receita'] * 100) if d['receita'] > 0 else 0

    top_produtos = sorted(por_produto.values(), key=lambda x: x['lucro'], reverse=True)[:15]

    lucro_total = receita_total - custo_total
    margem_pct = (lucro_total / receita_total * 100) if receita_total > 0 else 0
    ticket_lucro = (lucro_total / qtd_total) if qtd_total > 0 else 0

    # Itens sem snapshot de custo (lucro pode estar inflacionado)
    itens_sem_custo = qs.filter(preco_custo_unitario=0).count()

    return render(request, 'gestao/lucro.html', {
        'active_page': 'lucro',
        'periodo': periodo,
        'receita_total': receita_total,
        'custo_total': custo_total,
        'lucro_total': lucro_total,
        'margem_pct': margem_pct,
        'qtd_total': qtd_total,
        'ticket_lucro': ticket_lucro,
        'top_produtos': top_produtos,
        'itens_sem_custo': itens_sem_custo,
    })


# ---------------------------------------------------------------------------
# Promoções / Épocas Promocionais
# ---------------------------------------------------------------------------


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_promocoes(request):
    from datetime import date
    from .models import Cliente, Funcionario, ConfigAniversario, FelicitacaoEnviada
    hoje = date.today()
    promocoes = list(Promocao.objects.all())

    activas, futuras, passadas = [], [], []
    for p in promocoes:
        ini, fim = p.proxima_ocorrencia(hoje)
        p.ini = ini
        p.fim = fim
        p.a_decorrer = p.activo and p.esta_decorrer(hoje)
        p.dias_para_inicio = (ini - hoje).days
        p.dias_para_fim = (fim - hoje).days
        if not p.activo:
            passadas.append(p)
        elif p.a_decorrer:
            activas.append(p)
        else:
            futuras.append(p)

    futuras.sort(key=lambda x: x.dias_para_inicio)

    # ----- Aniversariantes do mês -----
    config = ConfigAniversario.get_solo()
    mes = hoje.month
    clientes_mes = list(
        Cliente.objects.filter(data_nascimento__month=mes).order_by('data_nascimento__day', 'nome')
    )
    funcionarios_mes = list(
        Funcionario.objects.filter(data_nascimento__month=mes, activo=True).order_by('data_nascimento__day', 'nome')
    )

    enviados_ano = set(
        FelicitacaoEnviada.objects.filter(ano=hoje.year, sucesso=True)
        .values_list('tipo', 'pessoa_id')
    )

    def _decorar(pessoa, tipo):
        d = pessoa.data_nascimento
        pessoa.dia = d.day
        pessoa.eh_hoje = (d.day == hoje.day and d.month == hoje.month)
        pessoa.eh_passado = (d.month == hoje.month and d.day < hoje.day)
        pessoa.dias_em_falta = max(0, d.day - hoje.day) if d.month == hoje.month else 0
        pessoa.idade = hoje.year - d.year - ((hoje.month, hoje.day) < (d.month, d.day))
        pessoa.tipo_pessoa = tipo
        pessoa.ja_felicitado = (tipo, pessoa.pk) in enviados_ano
        pessoa.preview = config.render_mensagem(pessoa, tipo)
        return pessoa

    clientes_mes = [_decorar(c, 'cliente') for c in clientes_mes]
    funcionarios_mes = [_decorar(f, 'funcionario') for f in funcionarios_mes]

    aniversariantes_hoje = [p for p in (clientes_mes + funcionarios_mes) if p.eh_hoje]

    return render(request, 'gestao/promocoes.html', {
        'active_page': 'promocoes',
        'activas': activas,
        'futuras': futuras,
        'passadas': passadas,
        'total': len(promocoes),
        'hoje': hoje,
        'config': config,
        'clientes_mes': clientes_mes,
        'funcionarios_mes': funcionarios_mes,
        'aniversariantes_hoje': aniversariantes_hoje,
        'mes_nome': ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                     'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes - 1],
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_promocao_criar(request, pk=None):
    instancia = get_object_or_404(Promocao, pk=pk) if pk else None

    if request.method == 'POST':
        from datetime import datetime
        nome = (request.POST.get('nome') or '').strip()
        tipo = request.POST.get('tipo') or 'dia'
        descricao = (request.POST.get('descricao') or '').strip()
        cor = request.POST.get('cor') or 'amber'
        icone = (request.POST.get('icone') or '🎁').strip()[:10]
        di = request.POST.get('data_inicio') or ''
        df = request.POST.get('data_fim') or ''
        try:
            desc_pct = int(request.POST.get('desconto_percentagem') or 0)
        except (ValueError, TypeError):
            desc_pct = 0
            erros.append('Percentagem de desconto inválida.')
        desc_pct = max(0, min(100, desc_pct))
        recorrente = request.POST.get('recorrente_anual') == 'on'
        activo = request.POST.get('activo') == 'on'
        subtitulo = (request.POST.get('subtitulo') or '').strip()[:200]
        texto_botao = (request.POST.get('texto_botao') or 'Ver Promoção').strip()[:50] or 'Ver Promoção'
        link_botao = (request.POST.get('link_botao') or '').strip()[:200]
        mostrar_carrossel = request.POST.get('mostrar_carrossel') == 'on'
        mostrar_landing = request.POST.get('mostrar_landing') == 'on'
        produto_ids = request.POST.getlist('produtos')
        remover_imagem = request.POST.get('remover_imagem') == 'on'
        nova_imagem = request.FILES.get('imagem')

        erros = []
        if not nome:
            erros.append('Nome é obrigatório.')
        try:
            data_inicio = datetime.strptime(di, '%Y-%m-%d').date()
        except ValueError:
            erros.append('Data de início inválida.')
            data_inicio = None
        try:
            data_fim = datetime.strptime(df, '%Y-%m-%d').date()
        except ValueError:
            erros.append('Data de fim inválida.')
            data_fim = None
        if data_inicio and data_fim and data_fim < data_inicio:
            erros.append('A data de fim deve ser posterior ou igual à de início.')

        if erros:
            for e in erros:
                messages.error(request, e)
        else:
            promo = instancia or Promocao()
            promo.nome = nome
            promo.tipo = tipo
            promo.descricao = descricao
            promo.cor = cor
            promo.icone = icone
            promo.data_inicio = data_inicio
            promo.data_fim = data_fim
            promo.desconto_percentagem = desc_pct
            promo.recorrente_anual = recorrente
            promo.activo = activo
            promo.subtitulo = subtitulo
            promo.texto_botao = texto_botao
            promo.link_botao = link_botao
            promo.mostrar_carrossel = mostrar_carrossel
            promo.mostrar_landing = mostrar_landing
            if remover_imagem and promo.imagem:
                promo.imagem.delete(save=False)
                promo.imagem = None
            if nova_imagem:
                promo.imagem = nova_imagem
            promo.save()
            try:
                ids_validos = [int(x) for x in produto_ids if str(x).isdigit()]
            except (TypeError, ValueError):
                ids_validos = []
            promo.produtos.set(Produto.objects.filter(pk__in=ids_validos))
            messages.success(request, f'Promoção "{nome}" {"actualizada" if instancia else "criada"} com sucesso.')
            return redirect('gestao_promocoes')

    if request.method == 'POST':
        valores = {
            'nome': request.POST.get('nome', ''),
            'tipo': request.POST.get('tipo', 'dia'),
            'descricao': request.POST.get('descricao', ''),
            'data_inicio': request.POST.get('data_inicio', ''),
            'data_fim': request.POST.get('data_fim', ''),
            'desconto_percentagem': request.POST.get('desconto_percentagem', '0'),
            'cor': request.POST.get('cor', 'amber'),
            'icone': request.POST.get('icone', '🎁'),
            'recorrente_anual': request.POST.get('recorrente_anual') == 'on',
            'activo': request.POST.get('activo') == 'on',
            'subtitulo': request.POST.get('subtitulo', ''),
            'texto_botao': request.POST.get('texto_botao', 'Ver Promoção'),
            'link_botao': request.POST.get('link_botao', ''),
            'mostrar_carrossel': request.POST.get('mostrar_carrossel') == 'on',
            'mostrar_landing': request.POST.get('mostrar_landing') == 'on',
        }
        produtos_selecionados_ids = [int(x) for x in request.POST.getlist('produtos') if str(x).isdigit()]
    elif instancia:
        valores = {
            'nome': instancia.nome,
            'tipo': instancia.tipo,
            'descricao': instancia.descricao,
            'data_inicio': instancia.data_inicio.strftime('%Y-%m-%d') if instancia.data_inicio else '',
            'data_fim': instancia.data_fim.strftime('%Y-%m-%d') if instancia.data_fim else '',
            'desconto_percentagem': instancia.desconto_percentagem,
            'cor': instancia.cor,
            'icone': instancia.icone or '🎁',
            'recorrente_anual': instancia.recorrente_anual,
            'activo': instancia.activo,
            'subtitulo': instancia.subtitulo,
            'texto_botao': instancia.texto_botao or 'Ver Promoção',
            'link_botao': instancia.link_botao,
            'mostrar_carrossel': instancia.mostrar_carrossel,
            'mostrar_landing': instancia.mostrar_landing,
        }
        produtos_selecionados_ids = list(instancia.produtos.values_list('id', flat=True))
    else:
        valores = {
            'nome': '', 'tipo': 'dia', 'descricao': '',
            'data_inicio': '', 'data_fim': '',
            'desconto_percentagem': 0, 'cor': 'amber', 'icone': '🎁',
            'recorrente_anual': True, 'activo': True,
            'subtitulo': '', 'texto_botao': 'Ver Promoção', 'link_botao': '',
            'mostrar_carrossel': False, 'mostrar_landing': False,
        }
        produtos_selecionados_ids = []

    produtos_disponiveis = Produto.objects.filter(disponivel=True).order_by('genero', 'nome')

    return render(request, 'gestao/promocao_form.html', {
        'active_page': 'promocoes',
        'instancia': instancia,
        'tipos': Promocao.TIPO_CHOICES,
        'cores': ['amber', 'rose', 'sky', 'emerald', 'violet', 'stone'],
        'v': valores,
        'produtos_disponiveis': produtos_disponiveis,
        'produtos_selecionados_ids': produtos_selecionados_ids,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_promocao_apagar(request, pk):
    promo = get_object_or_404(Promocao, pk=pk)
    if request.method == 'POST':
        nome = promo.nome
        promo.delete()
        messages.success(request, f'Promoção "{nome}" eliminada.')
    return redirect('gestao_promocoes')


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_promocao_toggle(request, pk):
    """Activa / desactiva uma promoção rapidamente."""
    promo = get_object_or_404(Promocao, pk=pk)
    promo.activo = not promo.activo
    promo.save(update_fields=['activo'])
    messages.success(request, f'Promoção "{promo.nome}" {"activada" if promo.activo else "desactivada"}.')
    return redirect('gestao_promocoes')


# ---------------------------------------------------------------------------
# Aniversariantes — configuração e envio
# ---------------------------------------------------------------------------

def _enviar_felicitacao(pessoa, tipo, config, canal_forcado=None):
    """Envia felicitação a uma pessoa. Devolve lista de FelicitacaoEnviada criados.

    canal_forcado: None envia pelos canais configurados; 'email', 'sms' ou 'manual' força um.
    """
    from datetime import date as _date
    from django.core.mail import send_mail
    from django.conf import settings as _settings
    from .models import FelicitacaoEnviada

    mensagem = config.render_mensagem(pessoa, tipo)
    assunto = 'Feliz Aniversário! 🎉 — Décent Privé'
    ano = _date.today().year
    resultados = []

    canais = []
    if canal_forcado:
        canais = [canal_forcado]
    else:
        if config.enviar_email and getattr(pessoa, 'email', ''):
            canais.append('email')
        if config.enviar_sms and getattr(pessoa, 'telefone', ''):
            canais.append('sms')

    if not canais:
        log = FelicitacaoEnviada.objects.create(
            tipo=tipo, pessoa_id=pessoa.pk, nome=pessoa.nome,
            canal='manual', sucesso=False,
            erro='Sem email nem telefone configurado.', ano=ano,
        )
        resultados.append(log)
        return resultados

    for canal in canais:
        sucesso, erro = True, ''
        try:
            if canal == 'email':
                if not getattr(pessoa, 'email', ''):
                    raise ValueError('Cliente/Funcionário sem email.')
                send_mail(
                    assunto, mensagem,
                    getattr(_settings, 'DEFAULT_FROM_EMAIL', 'decentprive@gmail.com'),
                    [pessoa.email], fail_silently=False,
                )
            elif canal == 'sms':
                gateway = getattr(_settings, 'SMS_GATEWAY_URL', '') or ''
                if not gateway:
                    logger.info('SMS (dev) para %s: %s', pessoa.telefone, mensagem[:100])
                else:
                    import requests as _req
                    payload = {
                        'to': pessoa.telefone, 'message': mensagem,
                        'channel': 'whatsapp',
                    }
                    headers = {}
                    token = getattr(_settings, 'SMS_GATEWAY_TOKEN', '')
                    if token:
                        headers['Authorization'] = f'Bearer {token}'
                    r = _req.post(gateway, json=payload, headers=headers, timeout=10)
                    r.raise_for_status()
            elif canal == 'manual':
                pass  # marcação manual de envio
        except Exception as exc:
            sucesso = False
            erro = str(exc)[:300]
            logger.exception('Falha ao enviar felicitação via %s para %s (pk=%s): %s', canal, pessoa.nome, pessoa.pk, exc)

        log = FelicitacaoEnviada.objects.create(
            tipo=tipo, pessoa_id=pessoa.pk, nome=pessoa.nome,
            canal=canal, sucesso=sucesso, erro=erro, ano=ano,
        )
        resultados.append(log)

    return resultados


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_aniversarios_config(request):
    from .models import ConfigAniversario
    config = ConfigAniversario.get_solo()
    if request.method == 'POST':
        from datetime import datetime as _dt
        tem_erros = False
        config.mensagem_cliente = (request.POST.get('mensagem_cliente') or '').strip() or config.mensagem_cliente
        config.mensagem_funcionario = (request.POST.get('mensagem_funcionario') or '').strip() or config.mensagem_funcionario
        hora_raw = (request.POST.get('hora_envio') or '10:00').strip()
        try:
            config.hora_envio = _dt.strptime(hora_raw, '%H:%M').time()
        except ValueError:
            messages.error(request, 'Hora de envio inválida.')
            tem_erros = True
        config.horario_aleatorio = request.POST.get('horario_aleatorio') == 'on'
        ji_raw = (request.POST.get('janela_inicio') or '09:00').strip()
        jf_raw = (request.POST.get('janela_fim') or '18:00').strip()
        try:
            config.janela_inicio = _dt.strptime(ji_raw, '%H:%M').time()
        except ValueError:
            messages.error(request, 'Início da janela inválido.')
            tem_erros = True
        try:
            config.janela_fim = _dt.strptime(jf_raw, '%H:%M').time()
        except ValueError:
            messages.error(request, 'Fim da janela inválido.')
            tem_erros = True
        if not tem_erros and config.janela_fim <= config.janela_inicio:
            messages.error(request, 'O fim da janela deve ser posterior ao início.')
            tem_erros = True
        config.enviar_email = request.POST.get('enviar_email') == 'on'
        config.enviar_sms = request.POST.get('enviar_sms') == 'on'
        config.brinde_activo = request.POST.get('brinde_activo') == 'on'
        config.brinde_descricao = (request.POST.get('brinde_descricao') or '').strip()
        config.activo = request.POST.get('activo') == 'on'
        if tem_erros:
            return render(request, 'gestao/aniversarios_config.html', {
                'active_page': 'promocoes',
                'config': config,
            })
        config.save()
        messages.success(request, 'Configuração de aniversários actualizada.')
        return redirect('gestao_aniversarios_config')

    return render(request, 'gestao/aniversarios_config.html', {
        'active_page': 'promocoes',
        'config': config,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_aniversariante_enviar(request, tipo, pk):
    """Envia felicitação manual a um cliente ou funcionário (botão 'Enviar agora')."""
    from .models import Cliente, Funcionario, ConfigAniversario
    if tipo == 'cliente':
        pessoa = get_object_or_404(Cliente, pk=pk)
    elif tipo == 'funcionario':
        pessoa = get_object_or_404(Funcionario, pk=pk)
    else:
        return redirect('gestao_promocoes')

    if request.method != 'POST':
        return redirect('gestao_promocoes')

    config = ConfigAniversario.get_solo()
    canal = request.POST.get('canal') or None
    resultados = _enviar_felicitacao(pessoa, tipo, config, canal_forcado=canal)
    sucessos = [r for r in resultados if r.sucesso]
    falhas = [r for r in resultados if not r.sucesso]
    if sucessos:
        canais_ok = ', '.join(sorted({r.get_canal_display() for r in sucessos}))
        messages.success(request, f'Felicitação enviada a {pessoa.nome} via {canais_ok}.')
    for f in falhas:
        messages.error(request, f'Falha ({f.get_canal_display()}) para {pessoa.nome}: {f.erro}')
    return redirect('gestao_promocoes')


# ---------------------------------------------------------------------------
# Gestão de Salários
# ---------------------------------------------------------------------------

TAXA_INSS_TRABALHADOR = 3   # contribuição do trabalhador (Angola)
TAXA_INSS_PATRONAL   = 8   # contribuição da entidade patronal (Angola)


def _calcular_irt(bruto_mensal):
    """IRT Angola — Tabela A (trabalhadores por conta de outrem).
    Cálculo progressivo por escalões mensais (Lei nº 18/14 e atualizações).
    INSS e subsídios NÃO entram na base tributável (dedução antes do IRT).
    """
    b = float(bruto_mensal)
    if b <= 70000:
        return 0.0
    escaloes = [
        (70000,         0,       0.00),
        (100000,        70000,   0.10),
        (150000,        100000,  0.13),
        (200000,        150000,  0.16),
        (300000,        200000,  0.18),
        (500000,        300000,  0.19),
        (float('inf'), 500000,  0.20),
    ]
    irt = 0.0
    for limite, inicio, taxa in escaloes:
        if b <= inicio:
            break
        tributavel = min(b, limite) - inicio
        irt += tributavel * taxa
    return round(irt, 2)


def _calcular_vendas_funcionario(funcionario, mes, ano):
    """Soma o total das vendas finalizadas de um funcionário num dado mês/ano."""
    if not funcionario.utilizador:
        return 0
    from datetime import date
    qs = ItemEncomenda.objects.filter(
        encomenda__vendido_por=funcionario.utilizador,
        encomenda__status='finalizada',
        encomenda__criado_em__year=ano,
        encomenda__criado_em__month=mes,
    )
    return float(sum(item.subtotal() for item in qs))


def _criar_pagamento(f, mes_p, ano_p, user):
    """Cria um PagamentoSalario para o funcionário f no mês/ano indicado."""
    vendas = _calcular_vendas_funcionario(f, mes_p, ano_p)
    pct = float(f.percentagem_comissao or 0) if f.cargo in Funcionario.CARGO_COM_COMISSAO else 0
    comissao = vendas * pct / 100
    bruto = float(f.salario) + comissao
    desconto_inss = round(bruto * TAXA_INSS_TRABALHADOR / 100, 2)
    desconto_irt = _calcular_irt(bruto)
    custo_inss_patronal = round(bruto * TAXA_INSS_PATRONAL / 100, 2)
    sub_alim = float(f.subsidio_alimentacao or 0)
    sub_transp = float(f.subsidio_transporte or 0)
    liquido = round(bruto - desconto_inss - desconto_irt + sub_alim + sub_transp, 2)
    return PagamentoSalario.objects.create(
        funcionario=f,
        mes=mes_p, ano=ano_p,
        salario_base=f.salario,
        percentagem_comissao=f.percentagem_comissao,
        total_vendas_mes=round(vendas, 2),
        comissao_valor=round(comissao, 2),
        salario_bruto=round(bruto, 2),
        taxa_inss=TAXA_INSS_TRABALHADOR,
        desconto_inss=desconto_inss,
        desconto_irt=desconto_irt,
        subsidio_alimentacao=round(sub_alim, 2),
        subsidio_transporte=round(sub_transp, 2),
        custo_inss_patronal=custo_inss_patronal,
        salario_liquido=liquido,
        processado_por=user,
    )


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_salarios(request):
    """Lista e processa o processamento salarial mensal."""
    hoje = timezone.localdate()
    mes_sel = int(request.GET.get('mes', hoje.month))
    ano_sel = int(request.GET.get('ano', hoje.year))

    if request.method == 'POST':
        acao = request.POST.get('acao')
        mes_p = int(request.POST.get('mes', mes_sel))
        ano_p = int(request.POST.get('ano', ano_sel))

        # ── Gerar processamento ──────────────────────────────────────────
        if acao == 'gerar':
            funcionarios = Funcionario.objects.filter(activo=True)
            criados = 0
            for f in funcionarios:
                if not f.salario:
                    continue
                if PagamentoSalario.objects.filter(funcionario=f, mes=mes_p, ano=ano_p).exists():
                    continue
                _criar_pagamento(f, mes_p, ano_p, request.user)
                criados += 1
            if criados:
                messages.success(request, f'Processamento gerado para {criados} funcionário(s) — {mes_p:02d}/{ano_p}.')
            else:
                messages.warning(request, 'Processamento já existente ou sem funcionários com salário definido.')
            return redirect(f'{request.path}?mes={mes_p}&ano={ano_p}')

        # ── Recalcular (apagar e regenerar) ─────────────────────────────
        if acao == 'recalcular':
            apagados, _ = PagamentoSalario.objects.filter(mes=mes_p, ano=ano_p).delete()
            funcionarios = Funcionario.objects.filter(activo=True)
            criados = 0
            for f in funcionarios:
                if not f.salario:
                    continue
                _criar_pagamento(f, mes_p, ano_p, request.user)
                criados += 1
            messages.success(request, f'Processamento recalculado: {criados} funcionário(s) — {mes_p:02d}/{ano_p}.')
            return redirect(f'{request.path}?mes={mes_p}&ano={ano_p}')

        # ── Marcar todos como pagos ──────────────────────────────────────
        if acao == 'marcar_todos_pagos':
            atualizados = PagamentoSalario.objects.filter(
                mes=mes_p, ano=ano_p, pago=False
            ).update(pago=True, data_pagamento=hoje)
            messages.success(request, f'{atualizados} salário(s) marcado(s) como pagos.')
            return redirect(f'{request.path}?mes={mes_p}&ano={ano_p}')

        # ── Marcar um como pago ──────────────────────────────────────────
        if acao == 'marcar_pago':
            pk = request.POST.get('pk')
            pag = get_object_or_404(PagamentoSalario, pk=pk)
            pag.pago = True
            pag.data_pagamento = hoje
            pag.save(update_fields=['pago', 'data_pagamento'])
            messages.success(request, f'Salário de {pag.funcionario.nome} marcado como pago.')
            return redirect(f'{request.path}?mes={mes_sel}&ano={ano_sel}')

    pagamentos = PagamentoSalario.objects.filter(
        mes=mes_sel, ano=ano_sel
    ).select_related('funcionario').order_by('funcionario__nome')

    total_bruto        = sum(float(p.salario_bruto)        for p in pagamentos)
    total_inss         = sum(float(p.desconto_inss)        for p in pagamentos)
    total_irt          = sum(float(p.desconto_irt)         for p in pagamentos)
    total_subsidios    = sum(float(p.subsidio_alimentacao) + float(p.subsidio_transporte) for p in pagamentos)
    total_liquido      = sum(float(p.salario_liquido)      for p in pagamentos)
    total_comissoes    = sum(float(p.comissao_valor)       for p in pagamentos)
    total_custo_patronal = sum(float(p.custo_inss_patronal) for p in pagamentos)

    # Funcionários activos sem salário definido (aviso)
    ids_processados = set(p.funcionario_id for p in pagamentos)
    funcionarios_sem_salario = Funcionario.objects.filter(
        activo=True, salario__isnull=True
    )
    funcionarios_nao_processados = Funcionario.objects.filter(
        activo=True
    ).exclude(
        pk__in=ids_processados
    ).exclude(salario__isnull=True)

    pendentes = sum(1 for p in pagamentos if not p.pago)
    anos = range(hoje.year - 2, hoje.year + 1)

    return render(request, 'gestao/salarios.html', {
        'active_page': 'salarios',
        'pagamentos': pagamentos,
        'mes_sel': mes_sel,
        'ano_sel': ano_sel,
        'anos': anos,
        'total_bruto': total_bruto,
        'total_inss': total_inss,
        'total_irt': total_irt,
        'total_subsidios': total_subsidios,
        'total_liquido': total_liquido,
        'total_comissoes': total_comissoes,
        'total_custo_patronal': total_custo_patronal,
        'custo_total_empresa': total_bruto + total_custo_patronal,
        'pendentes': pendentes,
        'funcionarios_sem_salario': funcionarios_sem_salario,
        'funcionarios_nao_processados': funcionarios_nao_processados,
        'meses': PagamentoSalario.MESES,
        'taxa_inss': TAXA_INSS_TRABALHADOR,
        'taxa_inss_patronal': TAXA_INSS_PATRONAL,
    })


# ---------------------------------------------------------------------------
# Exportação Excel — utilitário comum
# ---------------------------------------------------------------------------

def _excel_response(wb, filename):
    """Devolve um HttpResponse com o workbook Excel."""
    from django.http import HttpResponse
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        from django.contrib import messages as _msg
        # Fallback — não pode exportar
        raise ImportError('openpyxl não está instalado. Execute: pip install openpyxl')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _estilo_cabecalho(ws, row, cols):
    """Aplica estilo de cabeçalho às colunas da linha indicada."""
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill('solid', fgColor='1C1917')
    font = Font(color='FFFFFF', bold=True, size=10)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_salarios_exportar_excel(request):
    """Exporta o processamento salarial de um mês para Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl não instalado. Execute: pip install openpyxl')
        return redirect('gestao_salarios')

    hoje = timezone.localdate()
    mes = int(request.GET.get('mes', hoje.month))
    ano = int(request.GET.get('ano', hoje.year))
    nome_mes = dict(PagamentoSalario.MESES).get(mes, str(mes))

    pagamentos = PagamentoSalario.objects.filter(
        mes=mes, ano=ano
    ).select_related('funcionario').order_by('funcionario__nome')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Salários {mes:02d}-{ano}'

    # Título
    NCOLS = 14
    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    ws[f'A1'] = f'Décent Privé — Processamento Salarial | {nome_mes} {ano}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    ws['A2'] = (
        f'INSS Trabalhador: {TAXA_INSS_TRABALHADOR}%  |  INSS Patronal: {TAXA_INSS_PATRONAL}%  |'
        f'  IRT: Tabela A Progressiva  |  Gerado em: {hoje.strftime("%d/%m/%Y")}'
    )
    ws['A2'].alignment = Alignment(horizontal='center')

    # Cabeçalhos
    headers = [
        'Funcionário', 'Cargo', 'Sal. Base (Kz)', '% Comissão',
        'Vendas Mês (Kz)', 'Comissão (Kz)', 'Bruto (Kz)',
        f'INSS {TAXA_INSS_TRABALHADOR}% (Kz)', 'IRT (Kz)',
        'Sub. Alim. (Kz)', 'Sub. Transp. (Kz)',
        'Líquido (Kz)', f'INSS Patronal {TAXA_INSS_PATRONAL}% (Kz)', 'Situação'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    _estilo_cabecalho(ws, 4, NCOLS)
    ws.row_dimensions[4].height = 20

    # Dados
    fill_alt = PatternFill('solid', fgColor='F5F5F4')
    fill_liq = PatternFill('solid', fgColor='D1FAE5')  # verde suave — coluna líquido
    fmt_kz = '#,##0.00'
    for i, p in enumerate(pagamentos):
        row = 5 + i
        bg = fill_alt if i % 2 == 1 else None
        data = [
            p.funcionario.nome,
            p.funcionario.get_cargo_display(),
            float(p.salario_base),
            float(p.percentagem_comissao),
            float(p.total_vendas_mes),
            float(p.comissao_valor),
            float(p.salario_bruto),
            float(p.desconto_inss),
            float(p.desconto_irt),
            float(p.subsidio_alimentacao),
            float(p.subsidio_transporte),
            float(p.salario_liquido),
            float(p.custo_inss_patronal),
            'Pago' if p.pago else 'Pendente',
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if bg:
                cell.fill = bg
            if isinstance(val, float):
                cell.number_format = fmt_kz
        ws.cell(row=row, column=12).fill = fill_liq  # destaque líquido

    # Totais
    total_row = 5 + len(pagamentos)
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    totais_cols = {
        3: 'salario_base', 6: 'comissao_valor', 7: 'salario_bruto',
        8: 'desconto_inss', 9: 'desconto_irt',
        10: 'subsidio_alimentacao', 11: 'subsidio_transporte',
        12: 'salario_liquido', 13: 'custo_inss_patronal',
    }
    fill_total = PatternFill('solid', fgColor='E7E5E4')
    for col in range(1, NCOLS + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = fill_total
        cell.font = Font(bold=True)
    for col, attr in totais_cols.items():
        v = sum(float(getattr(p, attr)) for p in pagamentos)
        ws.cell(row=total_row, column=col, value=v).number_format = fmt_kz
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).fill = fill_total

    # Larguras
    widths = [28, 16, 16, 12, 18, 16, 16, 14, 14, 14, 14, 16, 18, 12]
    for col, width in zip(range(1, NCOLS + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = width

    return _excel_response(wb, f'salarios_{mes:02d}_{ano}.xlsx')


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_relatorio_exportar_excel(request):
    """Exporta o relatório de vendas/produtos para Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl não instalado. Execute: pip install openpyxl')
        return redirect('gestao_relatorio')

    from django.db.models import F, FloatField, ExpressionWrapper
    hoje = timezone.localdate()
    periodo = request.GET.get('periodo', 'mes')
    if periodo == 'hoje':
        data_inicio = hoje
        rotulo = 'Hoje'
    elif periodo == 'semana':
        data_inicio = hoje - timezone.timedelta(days=7)
        rotulo = 'Últimos 7 Dias'
    elif periodo == 'todos':
        data_inicio = None
        rotulo = 'Todo o Período'
    else:
        data_inicio = hoje.replace(day=1)
        rotulo = f'Mês {hoje.month:02d}/{hoje.year}'

    itens_qs = ItemEncomenda.objects.filter(encomenda__status='finalizada')
    if data_inicio:
        itens_qs = itens_qs.filter(encomenda__criado_em__date__gte=data_inicio)

    top_produtos = (
        itens_qs.values('nome_produto')
        .annotate(
            qty=Sum('quantidade'),
            receita=Sum(ExpressionWrapper(F('preco_unitario') * F('quantidade'), output_field=FloatField())),
        ).order_by('-qty')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Top Produtos'

    ws.merge_cells('A1:D1')
    ws['A1'] = f'Décent Privé — Relatório de Vendas | {rotulo}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Produto', 'Qtd. Vendida', 'Receita (Kz)', 'Receita %']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _estilo_cabecalho(ws, 3, len(headers))

    total_receita = sum(p['receita'] or 0 for p in top_produtos)
    fill_alt = PatternFill('solid', fgColor='F5F5F4')
    for i, p in enumerate(top_produtos):
        row = 4 + i
        receita = float(p['receita'] or 0)
        pct = (receita / total_receita * 100) if total_receita else 0
        ws.cell(row=row, column=1, value=p['nome_produto'])
        ws.cell(row=row, column=2, value=p['qty'])
        ws.cell(row=row, column=3, value=receita)
        ws.cell(row=row, column=4, value=round(pct, 1))
        if i % 2 == 1:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill_alt

    for col, width in zip(range(1, 5), [40, 14, 18, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width

    return _excel_response(wb, f'relatorio_{periodo}_{hoje}.xlsx')


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_caixa_exportar_excel(request):
    """Exporta os movimentos de caixa (entradas e saídas) para Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl não instalado. Execute: pip install openpyxl')
        return redirect('gestao_caixa')

    hoje = timezone.localdate()
    movimentos = MovimentoCaixa.objects.select_related('criado_por').order_by('-data')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Movimentos de Caixa'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'Décent Privé — Movimentos de Caixa | Exportado em {hoje.strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Data', 'Tipo', 'Categoria', 'Descrição', 'Valor (Kz)', 'Registado por']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _estilo_cabecalho(ws, 3, len(headers))

    fill_entrada = PatternFill('solid', fgColor='DCFCE7')
    fill_saida = PatternFill('solid', fgColor='FEE2E2')
    for i, m in enumerate(movimentos):
        row = 4 + i
        ws.cell(row=row, column=1, value=m.data.strftime('%d/%m/%Y') if m.data else '')
        ws.cell(row=row, column=2, value=m.get_tipo_display())
        ws.cell(row=row, column=3, value=m.get_categoria_display())
        ws.cell(row=row, column=4, value=m.descricao)
        ws.cell(row=row, column=5, value=float(m.valor))
        ws.cell(row=row, column=6, value=m.criado_por.get_full_name() if m.criado_por else '')
        fill = fill_entrada if m.tipo == 'entrada' else fill_saida
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill

    for col, width in zip(range(1, 7), [14, 12, 22, 38, 16, 22]):
        ws.column_dimensions[get_column_letter(col)].width = width

    return _excel_response(wb, f'caixa_{hoje}.xlsx')


# ---------------------------------------------------------------------------
# Newsletter
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Lotes de Importação
# ---------------------------------------------------------------------------

@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacoes(request):
    lotes = LoteImportacao.objects.prefetch_related('itens').all()
    return render(request, 'gestao/importacoes.html', {
        'active_page': 'importacoes',
        'lotes': lotes,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacao_criar(request, pk=None):
    lote = get_object_or_404(LoteImportacao, pk=pk) if pk else None

    if request.method == 'POST':
        referencia = request.POST.get('referencia', '').strip()
        data_encomenda = request.POST.get('data_encomenda', '').strip()
        cambio = request.POST.get('cambio', '').strip().replace(',', '.')
        custo_transporte = request.POST.get('custo_transporte_unidade', '').strip().replace(',', '.')
        status = request.POST.get('status', 'em_preparacao')
        notas = request.POST.get('notas', '').strip()

        erros = []
        if not referencia:
            erros.append('A referência é obrigatória.')
        if not data_encomenda:
            erros.append('A data da encomenda é obrigatória.')
        try:
            cambio_val = float(cambio)
            if cambio_val <= 0:
                raise ValueError
        except (ValueError, TypeError):
            erros.append('Câmbio inválido.')
            cambio_val = None
        try:
            transporte_val = float(custo_transporte)
            if transporte_val < 0:
                raise ValueError
        except (ValueError, TypeError):
            erros.append('Custo de transporte inválido.')
            transporte_val = None

        if erros:
            for e in erros:
                messages.error(request, e)
        else:
            if lote:
                lote.referencia = referencia
                lote.data_encomenda = data_encomenda
                lote.cambio = cambio_val
                lote.custo_transporte_unidade = transporte_val
                lote.status = status
                lote.notas = notas
                lote.save()
                messages.success(request, 'Lote actualizado com sucesso.')
            else:
                lote = LoteImportacao.objects.create(
                    referencia=referencia,
                    data_encomenda=data_encomenda,
                    cambio=cambio_val,
                    custo_transporte_unidade=transporte_val,
                    status=status,
                    notas=notas,
                    criado_por=request.user,
                )
                messages.success(request, 'Lote criado. Adicione os produtos abaixo.')
            return redirect('gestao_importacao_detalhe', pk=lote.pk)

    return render(request, 'gestao/importacao_form.html', {
        'active_page': 'importacoes',
        'lote': lote,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacao_detalhe(request, pk):
    lote = get_object_or_404(LoteImportacao.objects.prefetch_related('itens__produto'), pk=pk)
    produtos_catalogo = Produto.objects.filter(disponivel=True).order_by('nome')
    return render(request, 'gestao/importacao_detalhe.html', {
        'active_page': 'importacoes',
        'lote': lote,
        'produtos_catalogo': produtos_catalogo,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacao_apagar(request, pk):
    lote = get_object_or_404(LoteImportacao, pk=pk)
    if request.method == 'POST':
        ref = lote.referencia
        lote.delete()
        messages.success(request, f'Lote "{ref}" apagado.')
        return redirect('gestao_importacoes')
    return render(request, 'gestao/confirmar_apagar_lote.html', {
        'active_page': 'importacoes',
        'lote': lote,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacao_item_salvar(request, lote_pk, item_pk=None):
    lote = get_object_or_404(LoteImportacao, pk=lote_pk)
    item = get_object_or_404(ItemLoteImportacao, pk=item_pk, lote=lote) if item_pk else None

    if request.method != 'POST':
        return redirect('gestao_importacao_detalhe', pk=lote_pk)

    nome = request.POST.get('nome_produto', '').strip()
    marca = request.POST.get('marca', '').strip()
    genero = request.POST.get('genero', 'unissex')
    volume_ml = request.POST.get('volume_ml', '').strip()
    quantidade = request.POST.get('quantidade', '1').strip()
    preco_eur = request.POST.get('preco_compra_eur', '').strip().replace(',', '.')
    factor = request.POST.get('factor_lucro', '').strip().replace(',', '.')
    observacoes = request.POST.get('observacoes', '').strip()
    produto_id = request.POST.get('produto_id', '').strip()

    erros = []
    if not nome:
        erros.append('O nome do produto é obrigatório.')
    try:
        qty_val = int(quantidade)
        if qty_val <= 0:
            raise ValueError
    except (ValueError, TypeError):
        erros.append('Quantidade inválida.')
        qty_val = 1
    try:
        eur_val = float(preco_eur)
        if eur_val <= 0:
            raise ValueError
    except (ValueError, TypeError):
        erros.append('Preço de compra (€) inválido.')
        eur_val = None
    try:
        factor_val = float(factor)
        if not (0 < factor_val <= 1):
            raise ValueError
    except (ValueError, TypeError):
        erros.append('Factor de lucro inválido (deve estar entre 0 e 1, ex: 0.58).')
        factor_val = None

    produto_obj = None
    if produto_id:
        try:
            produto_obj = Produto.objects.get(pk=int(produto_id))
        except (Produto.DoesNotExist, ValueError):
            pass

    if erros:
        for e in erros:
            messages.error(request, e)
        return redirect('gestao_importacao_detalhe', pk=lote_pk)

    if item:
        item.nome_produto = nome
        item.marca = marca
        item.genero = genero
        item.volume_ml = volume_ml
        item.quantidade = qty_val
        item.preco_compra_eur = eur_val
        item.factor_lucro = factor_val
        item.observacoes = observacoes
        item.produto = produto_obj
        item.save()
        messages.success(request, f'"{nome}" actualizado.')
    else:
        ItemLoteImportacao.objects.create(
            lote=lote,
            nome_produto=nome,
            marca=marca,
            genero=genero,
            volume_ml=volume_ml,
            quantidade=qty_val,
            preco_compra_eur=eur_val,
            factor_lucro=factor_val,
            observacoes=observacoes,
            produto=produto_obj,
        )
        messages.success(request, f'"{nome}" adicionado ao lote.')

    return redirect('gestao_importacao_detalhe', pk=lote_pk)


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacao_item_apagar(request, lote_pk, item_pk):
    lote = get_object_or_404(LoteImportacao, pk=lote_pk)
    item = get_object_or_404(ItemLoteImportacao, pk=item_pk, lote=lote)
    if request.method == 'POST':
        nome = item.nome_produto
        item.delete()
        messages.success(request, f'"{nome}" removido do lote.')
    return redirect('gestao_importacao_detalhe', pk=lote_pk)


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacao_item_toggle(request, lote_pk, item_pk):
    """Toggle AJAX para marcar/desmarcar item como recebido."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido.'}, status=405)
    lote = get_object_or_404(LoteImportacao, pk=lote_pk)
    item = get_object_or_404(ItemLoteImportacao, pk=item_pk, lote=lote)
    item.recebido = not item.recebido
    item.save(update_fields=['recebido'])

    total = lote.itens.count()
    recebidos = lote.itens.filter(recebido=True).count()
    if recebidos == total and total > 0:
        novo_status = 'recebido'
    elif recebidos > 0:
        novo_status = 'parcial'
    else:
        novo_status = lote.status
    if novo_status != lote.status and novo_status in ('recebido', 'parcial'):
        lote.status = novo_status
        lote.save(update_fields=['status'])

    return JsonResponse({
        'recebido': item.recebido,
        'total': total,
        'recebidos': recebidos,
        'lote_status': lote.status,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_importacao_aplicar_precos(request, pk):
    """Aplica os preços calculados do lote aos produtos do catálogo ligados."""
    if request.method != 'POST':
        return redirect('gestao_importacao_detalhe', pk=pk)

    lote = get_object_or_404(LoteImportacao, pk=pk)
    actualizados = 0
    ignorados = 0

    for item in lote.itens.select_related('produto').all():
        if not item.produto:
            ignorados += 1
            continue
        pvp = round(item.preco_venda_sugerido, 2)
        custo = round(item.custo_unitario_kzs, 2)
        if pvp > 0:
            item.produto.preco_venda = pvp
            item.produto.preco_compra = custo
            item.produto.save(update_fields=['preco_venda', 'preco_compra'])
            actualizados += 1
        else:
            ignorados += 1

    if actualizados:
        msg = f'{actualizados} produto(s) actualizado(s) com os preços calculados.'
        if ignorados:
            msg += f' {ignorados} item(ns) sem ligação a produto foram ignorados.'
        messages.success(request, msg)
    else:
        messages.warning(request, 'Nenhum produto foi actualizado. Certifique-se de que os itens estão ligados a produtos do catálogo.')

    return redirect('gestao_importacao_detalhe', pk=pk)


# ---------------------------------------------------------------------------
# Gestão — Galeria de Imagens
# ---------------------------------------------------------------------------

@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_galeria(request):
    imagens = ImagemGaleria.objects.all()
    return render(request, 'gestao/galeria_imagens.html', {
        'active_page': 'galeria',
        'imagens': imagens,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_galeria_imagem_form(request, pk=None):
    instancia = get_object_or_404(ImagemGaleria, pk=pk) if pk else None
    if request.method == 'POST':
        legenda = request.POST.get('legenda', '').strip()[:200]
        try:
            ordem = int(request.POST.get('ordem') or 0)
        except (ValueError, TypeError):
            ordem = 0
        activo = request.POST.get('activo') == 'on'
        nova_imagem = request.FILES.get('imagem')
        remover_imagem = request.POST.get('remover_imagem') == 'on'

        if not instancia and not nova_imagem:
            messages.error(request, 'É obrigatório selecionar uma imagem.')
        else:
            obj = instancia or ImagemGaleria()
            obj.legenda = legenda
            obj.ordem = ordem
            obj.activo = activo
            if remover_imagem and obj.pk and obj.imagem:
                obj.imagem.delete(save=False)
                obj.imagem = None
            if nova_imagem:
                obj.imagem = nova_imagem
            obj.save()
            messages.success(request, 'Imagem guardada com sucesso.')
            return redirect('gestao_galeria')
    return render(request, 'gestao/galeria_imagem_form.html', {
        'active_page': 'galeria',
        'instancia': instancia,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_galeria_imagem_apagar(request, pk):
    img = get_object_or_404(ImagemGaleria, pk=pk)
    if request.method == 'POST':
        if img.imagem:
            img.imagem.delete(save=False)
        img.delete()
        messages.success(request, 'Imagem eliminada.')
    return redirect('gestao_galeria')


# ---------------------------------------------------------------------------
# Gestão — Banners / Carrossel
# ---------------------------------------------------------------------------

@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_banners(request):
    banners = Banner.objects.all()
    return render(request, 'gestao/banners.html', {
        'active_page': 'banners',
        'banners': banners,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_banner_form(request, pk=None):
    instancia = get_object_or_404(Banner, pk=pk) if pk else None
    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        subtitulo = request.POST.get('subtitulo', '').strip()[:300]
        texto_botao = (request.POST.get('texto_botao') or 'Ver Colecção').strip()[:50] or 'Ver Colecção'
        link_botao = (request.POST.get('link_botao') or '#produtos').strip()[:200]
        cor_fundo = (request.POST.get('cor_fundo') or '#1c1917').strip()[:7]
        try:
            ordem = int(request.POST.get('ordem') or 0)
        except (ValueError, TypeError):
            ordem = 0
        activo = request.POST.get('activo') == 'on'
        nova_imagem = request.FILES.get('imagem')
        remover_imagem = request.POST.get('remover_imagem') == 'on'

        if not titulo:
            messages.error(request, 'O título é obrigatório.')
        else:
            obj = instancia or Banner()
            obj.titulo = titulo
            obj.subtitulo = subtitulo
            obj.texto_botao = texto_botao
            obj.link_botao = link_botao
            obj.cor_fundo = cor_fundo
            obj.ordem = ordem
            obj.activo = activo
            if remover_imagem and obj.pk and obj.imagem:
                obj.imagem.delete(save=False)
                obj.imagem = None
            if nova_imagem:
                obj.imagem = nova_imagem
            obj.save()
            messages.success(request, 'Banner guardado com sucesso.')
            return redirect('gestao_banners')
    return render(request, 'gestao/banner_form.html', {
        'active_page': 'banners',
        'instancia': instancia,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_banner_apagar(request, pk):
    banner = get_object_or_404(Banner, pk=pk)
    if request.method == 'POST':
        if banner.imagem:
            banner.imagem.delete(save=False)
        banner.delete()
        messages.success(request, 'Banner eliminado.')
    return redirect('gestao_banners')


# ---------------------------------------------------------------------------
# Cupões de Desconto
# ---------------------------------------------------------------------------

@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_cupoes(request):
    cupoes = Cupao.objects.all().order_by('-criado_em')
    return render(request, 'gestao/cupoes.html', {
        'active_page': 'cupoes',
        'cupoes': cupoes,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_cupao_criar(request, pk=None):
    instancia = get_object_or_404(Cupao, pk=pk) if pk else None
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip().upper()
        descricao = request.POST.get('descricao', '').strip()
        tipo_desconto = request.POST.get('tipo_desconto', 'percentagem')
        valor_desconto = request.POST.get('valor_desconto', '0').strip()
        valor_minimo_compra = request.POST.get('valor_minimo_compra', '0').strip()
        limite_uso = request.POST.get('limite_uso', '0').strip()
        data_inicio = request.POST.get('data_inicio', '').strip()
        data_fim = request.POST.get('data_fim', '').strip()
        ativo = request.POST.get('ativo') == 'on'

        erros = {}
        if not codigo:
            erros['codigo'] = 'Código é obrigatório.'
        elif Cupao.objects.filter(codigo=codigo).exclude(pk=instancia.pk if instancia else None).exists():
            erros['codigo'] = 'Já existe um cupão com este código.'

        try:
            valor_desconto = float(valor_desconto)
            if valor_desconto <= 0:
                erros['valor_desconto'] = 'Valor deve ser maior que zero.'
            if tipo_desconto == 'percentagem' and valor_desconto > 100:
                erros['valor_desconto'] = 'Percentagem não pode exceder 100%.'
        except (ValueError, TypeError):
            erros['valor_desconto'] = 'Valor inválido.'

        try:
            valor_minimo_compra = float(valor_minimo_compra or 0)
            if valor_minimo_compra < 0:
                erros['valor_minimo_compra'] = 'Valor mínimo não pode ser negativo.'
        except (ValueError, TypeError):
            erros['valor_minimo_compra'] = 'Valor inválido.'

        try:
            limite_uso = int(limite_uso or 0)
            if limite_uso < 0:
                erros['limite_uso'] = 'Limite não pode ser negativo.'
        except (ValueError, TypeError):
            erros['limite_uso'] = 'Valor inválido.'

        if not data_inicio:
            erros['data_inicio'] = 'Data de início é obrigatória.'
        if not data_fim:
            erros['data_fim'] = 'Data de fim é obrigatória.'

        if erros:
            produtos = Produto.objects.filter(disponivel=True).order_by('nome')
            produtos_selecionados = set(int(x) for x in request.POST.getlist('produtos'))
            return render(request, 'gestao/cupao_form.html', {
                'active_page': 'cupoes',
                'instancia': instancia,
                'erros': erros,
                'post': request.POST,
                'produtos': produtos,
                'produtos_selecionados': produtos_selecionados,
            })

        from datetime import datetime
        dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%dT%H:%M')
        dt_fim = datetime.strptime(data_fim, '%Y-%m-%dT%H:%M')

        if dt_fim <= dt_inicio:
            produtos = Produto.objects.filter(disponivel=True).order_by('nome')
            produtos_selecionados = set(int(x) for x in request.POST.getlist('produtos'))
            return render(request, 'gestao/cupao_form.html', {
                'active_page': 'cupoes',
                'instancia': instancia,
                'erros': {'data_fim': 'Data de fim deve ser posterior à data de início.'},
                'post': request.POST,
                'produtos': produtos,
                'produtos_selecionados': produtos_selecionados,
            })

        obj = instancia or Cupao()
        obj.codigo = codigo
        obj.descricao = descricao
        obj.tipo_desconto = tipo_desconto
        obj.valor_desconto = valor_desconto
        obj.valor_minimo_compra = valor_minimo_compra
        obj.limite_uso = limite_uso
        obj.data_inicio = dt_inicio
        obj.data_fim = dt_fim
        obj.ativo = ativo
        obj.save()

        # Associar produtos (se seleccionados)
        produtos_ids = request.POST.getlist('produtos')
        if produtos_ids:
            obj.produtos.set(Produto.objects.filter(id__in=produtos_ids))
        else:
            obj.produtos.clear()

        messages.success(request, f'Cupão "{codigo}" guardado com sucesso.')
        return redirect('gestao_cupoes')

    produtos = Produto.objects.filter(disponivel=True).order_by('nome')
    produtos_selecionados = set(instancia.produtos.values_list('id', flat=True)) if instancia else set()
    return render(request, 'gestao/cupao_form.html', {
        'active_page': 'cupoes',
        'instancia': instancia,
        'produtos': produtos,
        'produtos_selecionados': produtos_selecionados,
    })


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_cupao_apagar(request, pk):
    cupao = get_object_or_404(Cupao, pk=pk)
    if request.method == 'POST':
        codigo = cupao.codigo
        cupao.delete()
        messages.success(request, f'Cupão "{codigo}" eliminado.')
    return redirect('gestao_cupoes')


@staff_member_required(login_url='/gestao/login/')
@_block_operador
def gestao_cupao_toggle(request, pk):
    if request.method != 'POST':
        return redirect('gestao_cupoes')
    cupao = get_object_or_404(Cupao, pk=pk)
    cupao.ativo = not cupao.ativo
    cupao.save(update_fields=['ativo'])
    messages.success(request, f'Cupão "{cupao.codigo}" {"activado" if cupao.ativo else "desactivado"}.')
    return redirect('gestao_cupoes')


@require_POST
def validar_cupao(request):
    """Endpoint AJAX para validar um cupão no checkout da loja."""
    import json
    data = json.loads(request.body or '{}')
    codigo = (data.get('codigo') or '').strip().upper()
    total_carrinho = float(data.get('total', 0))

    if not codigo:
        return JsonResponse({'ok': False, 'erro': 'Informe um código.'})

    try:
        cupao = Cupao.objects.get(codigo=codigo)
    except Cupao.DoesNotExist:
        return JsonResponse({'ok': False, 'erro': 'Cupão não encontrado.'})

    valido, msg = cupao.esta_valido()
    if not valido:
        return JsonResponse({'ok': False, 'erro': msg})

    if float(cupao.valor_minimo_compra) > 0 and total_carrinho < float(cupao.valor_minimo_compra):
        return JsonResponse({
            'ok': False,
            'erro': f'Compra mínima de {float(cupao.valor_minimo_compra):,.0f} Kz para este cupão.'
        })

    desconto = cupao.calcular_desconto(total_carrinho, itens=data.get('itens'))
    if desconto == 0 and cupao.produtos.exists():
        return JsonResponse({
            'ok': False,
            'erro': 'Este cupão só se aplica a produtos específicos que não estão no carrinho.'
        })
    return JsonResponse({
        'ok': True,
        'desconto': round(desconto, 2),
        'total_final': round(total_carrinho - desconto, 2),
        'tipo': cupao.tipo_desconto,
        'valor': float(cupao.valor_desconto),
        'codigo': cupao.codigo,
    })
